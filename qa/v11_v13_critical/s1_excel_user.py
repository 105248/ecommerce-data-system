# -*- coding: utf-8 -*-
"""SP02(Excel原值=Core抽样) + SP08(跨店人数语义) 快速检查
SP02: 读 raw_files 成交概览 Excel(6/30 官方) 关键字段 vs core.douyin_deal_daily；openpyxl read_only=False"""
import sys
sys.path.insert(0, r"D:/ecommerce-data-system/mcp_server")
from pathlib import Path
import psycopg2
from psycopg2.extras import RealDictCursor

# .env 密码
_env = {}
for _l in Path(r"D:/ecommerce-data-system/mcp_server/.env").read_text(encoding="utf-8").splitlines():
    _l = _l.strip()
    if _l and "=" in _l and not _l.startswith("#"):
        _k, _, _v = _l.partition("="); _env[_k.strip()] = _v.strip()
conn = psycopg2.connect(host="127.0.0.1", port=5432, dbname="ecommerce_db",
                        user="postgres", password=_env.get("PG_ADMIN_PASSWORD", ""), connect_timeout=5)
conn.set_session(readonly=True, autocommit=True)
cur = conn.cursor(cursor_factory=RealDictCursor)
findings = []
def rec(level, sp, desc):
    findings.append((level, sp, desc))
    print("[{}] {}: {}".format(level, sp, desc[:120]))

# ===== SP02 Excel 原值 = Core =====
print("\n===== SP02 Excel vs Core =====")
import openpyxl
xl_path = Path(r"D:/ecommerce-data-system/raw_files/douyin/弹动官方旗舰店/2026/2026-06/抖音电商罗盘-成交分析-20260601-20260630.xlsx")
if not xl_path.exists():
    rec("WARN", "SP02", "Excel 不存在: {}".format(xl_path))
else:
    wb = openpyxl.load_workbook(xl_path, read_only=False, data_only=True)
    print("sheets:", wb.sheetnames[:8])
    ws = wb["成交概览"]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h else "" for h in rows[0]]
    print("列数:", len(header), "数据行:", len(rows) - 2)
    # 找 2026-06-30 行
    def find_col(names):
        for i, h in enumerate(header):
            if any(n in h for n in names):
                return i
        return None
    # 关键字段列定位
    cols = {
        "结算金额": find_col(["结算金额"]),
        "投放消耗(店铺被投)": find_col(["投放消耗"]) if find_col(["投放消耗"]) is not None else None,
        "成交退款金额(支付时间)": find_col(["成交退款金额", "退款金额"]),
        "投放费比(剔除退款、店铺绑定)": find_col(["投放费比", "费比"]),
    }
    print("列定位:", {k: (header[v] if v is not None else None) for k, v in cols.items()})
    # 日期列
    date_col = None
    for i, h in enumerate(header):
        if h in ("日期", "数据日期"):
            date_col = i; break
    # 抽 2026-06-30 的"全店"行：日期+销售范围=全部+投放时段=不限
    target_date = "20260630"
    sample = None
    for r in rows[2:]:
        d = str(r[date_col]).replace("-", "").strip() if date_col is not None and r[date_col] else ""
        sc = str(r[1]).strip() if len(r) > 1 and r[1] else ""
        ap = str(r[2]).strip() if len(r) > 2 and r[2] else ""
        if d == target_date and sc == "全部" and ap == "不限":
            sample = r
            break
    if sample is None:
        # 回退：仅日期匹配第一个
        for r in rows[2:]:
            d = str(r[date_col]).replace("-", "").strip() if r[date_col] else ""
            if d == target_date:
                sample = r
                rec("WARN", "SP02", "未找到 全部/不限 行，用首个 6/30 行（第2列={} 第3列={}）".format(
                    r[1] if len(r) > 1 else "?", r[2] if len(r) > 2 else "?"))
                break
    if sample is None:
        rec("WARN", "SP02", "未找到 2026-06-30 行（前3行日期: {}）".format([str(r[date_col])[:10] if r[date_col] else "?" for r in rows[2:5]]))
    else:
        print("匹配行前3列:", [str(v)[:20] for v in sample[:4]])
        # 与 core 对照（deal_daily 官方店 6/30 全店）
        cur.execute("""SELECT settlement_amount, ad_spend_shop_promoted, ad_spend_shop_bound,
                              refund_amount_pay_time, transaction_refund_amount_pay_time,
                              ad_spend_rate_net_refund_shop_bound, total_expense_rate_net_refund_shop_bound
                       FROM core.douyin_deal_daily
                       WHERE shop_id=1 AND biz_date='2026-06-30' AND ad_period='不限'
                             AND sale_scope='全部' AND carrier_type='全部'""")
        core_row = cur.fetchone()
        if core_row:
            for cname, cidx in cols.items():
                if cidx is None:
                    continue
                xl_v = sample[cidx]
                core_v = None
                cmap = {"结算金额": "settlement_amount", "成交退款金额(支付时间)": "refund_amount_pay_time",
                        "投放消耗(店铺被投)": "ad_spend_shop_promoted"}
                if cname in cmap:
                    core_v = core_row[cmap[cname]]
                try:
                    xl_f = float(xl_v) if xl_v not in (None, "") else None
                    c_f = float(core_v) if core_v is not None else None
                    if xl_f is not None and c_f is not None:
                        diff = abs(xl_f - c_f)
                        ok = diff < 0.01
                        rec("PASS" if ok else "FAIL", "SP02", "{}: Excel={} Core={} diff={:.4f}".format(cname, xl_f, c_f, diff))
                except Exception as e:
                    rec("WARN", "SP02", "{}: 转换异常 {}".format(cname, str(e)[:40]))
        else:
            rec("WARN", "SP02", "core 无 6/30 全店行（ad_period 结构可能不同）")
    wb.close()

# ===== SP08 跨店人数语义 =====
print("\n===== SP08 跨店人数语义 =====")
# 1. 平台函数 notes（已确认）
# 2. 字段中文名（field_mapping/whitelist）
cur.execute("""SELECT DISTINCT target_column_name, target_column_name_cn FROM meta.field_mapping
               WHERE target_column_name LIKE '%buyer%' OR target_column_name_cn LIKE '%人数%' OR target_column_name_cn LIKE '%买家%'
               ORDER BY 1""")
buyer_fields = cur.fetchall()
rec("INFO", "SP08", "人数相关字段: {}".format([(r["target_column_name"], r["target_column_name_cn"]) for r in buyer_fields][:8]))
# 3. MCP/AI 描述是否用"唯一"
import re
hits = []
for f in ["platform_tools.py", "business_tools.py"]:
    txt = (Path(r"D:/ecommerce-data-system/mcp_server/tools") / f).read_text(encoding="utf-8")
    for i, line in enumerate(txt.splitlines(), 1):
        if "唯一人数" in line or "去重" in line or "unique" in line.lower():
            hits.append("{}:{}".format(f, i))
try:
    txt = (Path(r"D:/ecommerce-data-system/mcp_server/schemas.py")).read_text(encoding="utf-8")
    for i, line in enumerate(txt.splitlines(), 1):
        if "唯一人数" in line or "去重" in line or "unique" in line.lower():
            hits.append("schemas.py:{}".format(i))
except Exception:
    pass
rec("PASS" if not hits else "WARN", "SP08", "代码中'唯一人数/去重'表述: {}".format(hits or "无（平台层已声明'各店之和，跨店不去重'）"))
# AI prompt
try:
    p = Path(r"D:/ecommerce-data-system/ai_layer/system_prompt.md")
    if p.exists():
        txt = p.read_text(encoding="utf-8")
        rec("INFO", "SP08", "AI prompt 含'人数'相关: {}".format(
            ["唯一" in line or "去重" in line for line in txt.splitlines() if "人数" in line][:5] or "无唯一表述"))
except Exception as e:
    rec("WARN", "SP08", "AI prompt: {}".format(str(e)[:40]))
conn.close()
print("\n===== SP02/SP08 完成，发现 {} 条 =====".format(len(findings)))
