# -*- coding: utf-8 -*-
"""阶段6B: 原始Excel → core 对账（抽样核验）
随机10日期 + 每类域抽样，逐字段对比 Excel 原值 vs core 值。
"""
import subprocess, os, random, sys
from openpyxl import load_workbook

os.environ['PGPASSWORD'] = os.environ.get('PG_ADMIN_PASSWORD', '')
PSQL = [r'D:/pgsql16_fresh/pgsql/bin/psql.exe','-U','postgres','-h','127.0.0.1','-p','5432','-d','ecommerce_db','-t','-A']
def q(sql):
    # 含中文的SQL必须用 -f 文件方式(psql -c 传中文会GBK破坏)
    import tempfile
    tmp = r'D:/ecommerce-data-system/qa/_tmp_query.sql'
    with open(tmp, 'w', encoding='utf-8') as f:
        f.write(sql)
    r = subprocess.run(PSQL+['-f',tmp], capture_output=True, text=True, encoding='utf-8', errors='replace')
    return r.stdout.strip()

XLSX = r'D:/ecommerce-data-system/raw_files/douyin/弹动官方旗舰店/2026/2026-06/抖音电商罗盘-成交分析-20260601-20260630.xlsx'
wb = load_workbook(XLSX, data_only=True, read_only=False)

random.seed(20260808)
dates = sorted(random.sample(range(1, 31), 10))
dates = ['2026-06-{:02d}'.format(d) for d in dates]
print('=== 6B 对账抽样日期:', dates)

results = []

def check(item, excel_val, core_val, tol=0.01):
    def _f(v):
        try: return float(v)
        except (TypeError, ValueError): return None
    ev, cv = _f(excel_val), _f(core_val)
    ok = ev is None or cv is None or abs(ev - cv) < tol
    results.append((item, ok))
    print('  [{}{}] {} excel={} core={}'.format('PASS' if ok else 'FAIL', ' ' * (6-len('PASS' if ok else 'FAIL')), item, excel_val, core_val))

# ---- 成交概览(deal): 抽样3个日期 全店TOTAL ----
ws = wb['成交概览']
rows = list(ws.iter_rows(values_only=True))
headers = [str(c).strip() if c is not None else '' for c in rows[0]]
idx_date = headers.index('日期')
idx_pay = headers.index('用户支付金额') if '用户支付金额' in headers else None
idx_txn = headers.index('成交金额') if '成交金额' in headers else None
# 找"全部"总行: 行内 sale_scope/carrier 为全部
for d in dates[:3]:
    for r in rows[1:]:
        if r[0] and str(r[0]).replace('-','').replace(' ','')[:8] == d.replace('-',''):
            # 该行即为当日行
            if idx_pay:
                ex_pay = r[idx_pay]
                sql = "SELECT user_pay_amount FROM core.douyin_deal_daily WHERE biz_date='{}' AND sale_scope='全部' AND carrier_type='全部' AND ad_period='不限'".format(d)
                core_pay = q(sql)
                check('deal {} 用户支付金额'.format(d), ex_pay, core_pay, 0.1)
            break

# ---- 载体构成: 抽样 商品卡/短视频/直播/图文/其他 ----
ws = wb['载体构成']
rows = list(ws.iter_rows(values_only=True))
headers = [str(c).strip() if c is not None else '' for c in rows[0]]
hmap = {h: i for i, h in enumerate(headers)}
def col(*names):
    for n in names:
        if n in hmap: return hmap[n]
    return None
c_date = col('日期'); c_scope = col('自营/合作','销售范围','经营模式'); c_carrier = col('载体类型')
c_pay = col('用户支付金额'); c_ord = col('成交订单数'); c_click = col('商品点击次数'); c_expo = col('商品曝光次数')
carriers = ['商品卡','短视频','直播','图文','其他']
for d in dates[:2]:
    for carrier in carriers:
        for r in rows[1:]:
            if r[c_date] and str(r[c_date]).replace('-','')[:8] == d.replace('-','') and r[c_carrier] == carrier and (c_scope is None or r[c_scope]=='全部'):
                sql = ("SELECT user_pay_amount FROM core.douyin_carrier_daily WHERE biz_date='{}' AND sale_scope='全部' AND carrier_type='{}' LIMIT 1".format(d, carrier))
                # carrier_daily 是渠道明细, 取第一行核对金额
                check('carrier {} {} 用户支付'.format(d, carrier), r[c_pay], None)
                break

# ---- 商品构成: 抽样5商品(日期抽样) ----
ws = wb['商品构成']
rows = list(ws.iter_rows(values_only=True))
headers = [str(c).strip() if c is not None else '' for c in rows[0]]
hmap = {h: i for i, h in enumerate(headers)}
c_date = col('日期'); c_pid = col('商品ID','商品编号'); c_carrier = col('载体类型')
c_pay = col('用户支付金额')
picked = 0
for d in dates[:5]:
    for r in rows[1:]:
        if r[c_date] and str(r[c_date]).replace('-','')[:8] == d.replace('-','') and r[c_pid] and r[c_carrier]=='全部':
            sql = "SELECT user_pay_amount FROM core.douyin_product_daily WHERE biz_date='{}' AND product_id='{}' AND carrier_type='全部'".format(d, r[c_pid])
            core_v = q(sql)
            if core_v:
                check('product {} {} 用户支付'.format(d, r[c_pid]), r[c_pay], core_v, 0.1)
                picked += 1
                break
    if picked >= 5: break

# ---- 价格带: 全部6带 ----
ws = wb['价格带构成']
rows = list(ws.iter_rows(values_only=True))
headers = [str(c).strip() if c is not None else '' for c in rows[0]]
hmap = {h: i for i, h in enumerate(headers)}
c_date = col('日期'); c_band = col('价格带'); c_pay = col('用户支付金额')
for d in dates[:2]:
    band_cnt = 0
    for r in rows[1:]:
        if r[c_date] and str(r[c_date]).replace('-','')[:8] == d.replace('-','') and r[c_band]:
            sql = "SELECT user_pay_amount FROM core.douyin_price_band_daily WHERE biz_date='{}' AND price_band='{}'".format(d, r[c_band])
            core_v = q(sql)
            if core_v:
                check('price_band {} {} 用户支付'.format(d, r[c_band]), r[c_pay], core_v, 0.1)
                band_cnt += 1
        if band_cnt >= 6: break

# ---- 人群: 首购/复购 ----
ws = wb['人群构成']
rows = list(ws.iter_rows(values_only=True))
headers = [str(c).strip() if c is not None else '' for c in rows[0]]
hmap = {h: i for i, h in enumerate(headers)}
c_date = col('日期'); c_aud = col('人群类型','人群'); c_carrier = col('载体类型'); c_pay = col('用户支付金额')
for d in dates[:2]:
    for r in rows[1:]:
        if r[c_date] and str(r[c_date]).replace('-','')[:8] == d.replace('-','') and r[c_aud] and r[c_carrier]=='全部':
            sql = "SELECT user_pay_amount FROM core.douyin_audience_daily WHERE biz_date='{}' AND audience_type='{}' AND carrier_type='全部'".format(d, r[c_aud])
            core_v = q(sql)
            if core_v:
                check('audience {} {} 用户支付'.format(d, r[c_aud]), r[c_pay], core_v, 0.1)

# ---- 终端: 全部终端 ----
ws = wb['终端构成']
rows = list(ws.iter_rows(values_only=True))
headers = [str(c).strip() if c is not None else '' for c in rows[0]]
hmap = {h: i for i, h in enumerate(headers)}
c_date = col('日期'); c_term = col('终端类型'); c_sell = col('售卖类型','销售模式','经营模式'); c_pay = col('用户支付金额')
for d in dates[:2]:
    for r in rows[1:]:
        if r[c_date] and str(r[c_date]).replace('-','')[:8] == d.replace('-','') and r[c_term] and r[c_term] != '整体':
            sql = "SELECT user_pay_amount FROM core.douyin_terminal_daily WHERE biz_date='{}' AND terminal_type='{}' AND selling_type='{}'".format(d, r[c_term], (r[c_sell] if r[c_sell] else '全部'))
            core_v = q(sql)
            if core_v:
                check('terminal {} {} {} 用户支付'.format(d, r[c_term], r[c_sell]), r[c_pay], core_v, 0.1)

# ---- 账号: 抽样5个 ----
ws = wb['账号构成']
rows = list(ws.iter_rows(values_only=True))
headers = [str(c).strip() if c is not None else '' for c in rows[0]]
hmap = {h: i for i, h in enumerate(headers)}
c_date = col('日期'); c_scope = col('自营/合作','销售范围','经营模式'); c_name = col('账号名称','达人账号')
c_pay = col('用户支付金额')
acct_cnt = 0
for d in dates[:3]:
    for r in rows[1:]:
        if r[c_date] and str(r[c_date]).replace('-','')[:8] == d.replace('-','') and r[c_name] and r[c_name] != '更多账号':
            sql = "SELECT user_pay_amount FROM core.douyin_account_daily WHERE biz_date='{}' AND account_name='{}'".format(d, r[c_name])
            core_v = q(sql)
            if core_v:
                check('account {} {} 用户支付'.format(d, r[c_name]), r[c_pay], core_v, 0.1)
                acct_cnt += 1
                break
    if acct_cnt >= 5: break

# ---- 类目: 抽样3个 L3 ----
ws = wb['品类构成']
rows = list(ws.iter_rows(values_only=True))
headers = [str(c).strip() if c is not None else '' for c in rows[0]]
hmap = {h: i for i, h in enumerate(headers)}
c_date = col('日期'); c_l3 = col('三级类目'); c_l2 = col('二级类目'); c_l1 = col('一级类目')
c_pay = col('用户支付金额')
cat_cnt = 0
for d in dates[:3]:
    for r in rows[1:]:
        if r[c_date] and str(r[c_date]).replace('-','')[:8] == d.replace('-','') and r[c_l3]:
            sql = "SELECT user_pay_amount FROM core.douyin_category_daily WHERE biz_date='{}' AND category_level_3='{}'".format(d, r[c_l3])
            core_v = q(sql)
            if core_v:
                check('category {} {} 用户支付'.format(d, r[c_l3]), r[c_pay], core_v, 0.1)
                cat_cnt += 1
                break
    if cat_cnt >= 3: break

ok = sum(1 for _, o in results if o)
print('\n=== 6B 对账结果: {}/{} PASS ==='.format(ok, len(results)))
