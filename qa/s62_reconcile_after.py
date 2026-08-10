# v2: 覆盖后对账 (适配新文件结构: 成交概览=全部口径)
import subprocess, os
from openpyxl import load_workbook

os.environ['PGPASSWORD'] = os.environ.get('PG_ADMIN_PASSWORD', '')
PSQL = [r'D:/pgsql16_fresh/pgsql/bin/psql.exe','-U','postgres','-h','127.0.0.1','-p','5432','-d','ecommerce_db','-t','-A']
TMP = r'D:/ecommerce-data-system/qa/_tmp_s62q.sql'
def q(sql):
    with open(TMP, 'w', encoding='utf-8') as f:
        f.write(sql)
    r = subprocess.run(PSQL+['-f',TMP], capture_output=True, text=True, encoding='utf-8', errors='replace')
    return r.stdout.strip()

def dstr(v):
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    t = str(v).strip()
    if len(t) == 8 and t.isdigit():
        return t[0:4] + '-' + t[4:6] + '-' + t[6:8]
    return t[:10]

NEW = r'C:/Users/EDY/Downloads/抖音电商罗盘-成交分析-20260601-20260630 (3).xlsx'
wb = load_workbook(NEW, data_only=True, read_only=False)

results = []
def check(name, excel_val, core_val, tol=0.01):
    ok = (excel_val is None and core_val in ('', '0')) or (excel_val is not None and core_val not in ('', 'None') and abs(float(excel_val) - float(core_val)) < tol)
    results.append((name, ok))
    print(f'  [{"PASS" if ok else "FAIL"}] {name}: Excel={excel_val} core={core_val}')

# 1) 成交概览 全店 3天 (sale_scope=全部)
print('=== 1. 成交概览(全部口径) 3天 TOTAL ===')
ws = wb['成交概览']
rows = list(ws.iter_rows(values_only=True))
hdr = [str(c).strip() if c is not None else '' for c in rows[0]]
i_date, i_car, i_period, i_up = hdr.index('日期'), hdr.index('载体类型'), hdr.index('投放时段'), hdr.index('用户支付金额')
for r in rows[1:]:
    if dstr(r[i_date]) in ('2026-06-01','2026-06-15','2026-06-30') and r[i_car]=='全部' and r[i_period]=='不限':
        core = q(f"SELECT user_pay_amount FROM core.douyin_deal_daily WHERE biz_date='{r[i_date]}' AND sale_scope='全部' AND carrier_type='全部' AND ad_period='不限';")
        check(f'deal 全部 {dstr(r[i_date])}', r[i_up], core)

# 2) 自营成交 + 合作成交 total 3天
print()
print('=== 2. 自营/合作 total 3天 ===')
for sh, scope in [('自营成交','自营'), ('合作成交','合作')]:
    ws = wb[sh]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() if c is not None else '' for c in rows[0]]
    i_date, i_car, i_period, i_up = hdr.index('日期'), hdr.index('载体类型'), hdr.index('投放时段'), hdr.index('用户支付金额')
    for r in rows[1:]:
        if dstr(r[i_date]) in ('2026-06-05','2026-06-30') and r[i_car]=='全部' and r[i_period]=='不限':
            core = q(f"SELECT user_pay_amount FROM core.douyin_deal_daily WHERE biz_date='{r[i_date]}' AND sale_scope='{scope}' AND carrier_type='全部' AND ad_period='不限';")
            check(f'deal {scope} {dstr(r[i_date])}', r[i_up], core)

# 3) price_band 6带求和
print()
print('=== 3. 价格带 6带SUM ===')
ws = wb['价格带构成']
rows = list(ws.iter_rows(values_only=True))
hdr = [str(c).strip() if c is not None else '' for c in rows[0]]
i_up = hdr.index('用户支付金额')
xl_sum = sum(r[i_up] for r in rows[1:] if r[i_up] is not None)
core_sum = q("SELECT sum(user_pay_amount) FROM core.douyin_price_band_daily;")
check('price_band 6带SUM', xl_sum, core_sum)

# 4) carrier 剔除退款费比公式反算 (新数据)
print()
print('=== 4. carrier 剔除退款投放费比(被投) 公式反算 06-05 ===')
ws = wb['载体构成']
rows = list(ws.iter_rows(values_only=True))
hdr = [str(c).strip() if c is not None else '' for c in rows[0]]
i_date, i_scope = hdr.index('日期'), hdr.index('自营/合作')
i_asp = hdr.index('投放消耗(店铺被投)')
i_settle = hdr.index('结算金额')
i_rate = hdr.index('投放费比(剔除退款、店铺被投)')
ok_cnt = total = 0
for r in rows[1:]:
    if dstr(r[i_date]) != '2026-06-05' or not r[i_scope]:
        continue
    if r[i_asp] is None or r[i_settle] is None or r[i_rate] is None or r[i_settle] == 0:
        continue
    total += 1
    calc = r[i_asp] / r[i_settle]
    if abs(calc - r[i_rate]) < 0.0005:
        ok_cnt += 1
print(f'  06-05 剔除退款投放费比(被投) 公式成立: {ok_cnt}/{total}')

# 5) 账号 弹动官方旗舰店 06-05
print()
print('=== 5. 账号 弹动官方旗舰店 06-05 ===')
ws = wb['账号构成']
rows = list(ws.iter_rows(values_only=True))
hdr = [str(c).strip() if c is not None else '' for c in rows[0]]
i_date, i_name = hdr.index('日期'), hdr.index('账号名称')
i_up = hdr.index('用户支付金额')
for r in rows[1:]:
    if dstr(r[i_date]) == '2026-06-05' and r[i_name] == '弹动官方旗舰店':
        core = q("SELECT user_pay_amount FROM core.douyin_account_daily WHERE biz_date='2026-06-05' AND account_name='弹动官方旗舰店' LIMIT 1;")
        check('account 弹动官方旗舰店 06-05', r[i_up], core)
        break

print()
ok_all = sum(1 for x in results if x[1])
print(f'=== 对账汇总: {ok_all}/{len(results)} PASS ===')
for n, ok in results:
    if not ok:
        print(f'  ❌ {n}')
