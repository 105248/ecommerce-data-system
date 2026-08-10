# -*- coding: utf-8 -*-
"""F1.0.3 source inventory v2：统一 20 字段行结构（与 CSV header 完全对齐）"""
import csv
import hashlib
import re
import openpyxl
from pathlib import Path

CANDIDATES = [
    (r"C:/Users/EDY/Downloads/商品卡列表数据_2026_05_01~2026_05_31.xlsx", "商品卡列表", "官方", "douyin"),
    (r"C:/Users/EDY/Downloads/商品卡列表数据_2026_05_01~2026_05_31 (1).xlsx", "商品卡列表", "官方", "douyin"),
    (r"C:/Users/EDY/Downloads/商品卡列表数据_2026_06_01~2026_06_30.xlsx", "商品卡列表", "官方", "douyin"),
    (r"C:/Users/EDY/Downloads/商品卡列表数据_2026_06_01~2026_06_30 (1).xlsx", "商品卡列表", "官方", "douyin"),
    (r"C:/Users/EDY/Downloads/商品卡列表数据_2026_06_16~2026_07_15.xlsx", "商品卡列表", "官方", "douyin"),
    (r"C:/Users/EDY/Downloads/商品卡列表数据_2026_06_30~2026_07_06.xlsx", "商品卡列表", "官方", "douyin"),
    (r"C:/Users/EDY/Downloads/商品卡流量分析流量来源数据_2026_06_30~2026_07_06.xlsx", "商品卡流量来源", "官方", "douyin"),
    (r"C:/Users/EDY/Downloads/[20260630-20260706]_{自营}_{挂车}_{全部}_{所有账号}_{发布时间不限}视频详情数据.xlsx", "视频详情-自营挂车", "官方", "douyin"),
    (r"C:/Users/EDY/Downloads/[20260630-20260706]_{自营}_{非挂车}_{全部}_{所有账号}_{发布时间不限}视频详情数据.xlsx", "视频详情-自营非挂车", "官方", "douyin"),
    (r"C:/Users/EDY/Downloads/[20260630-20260706]_{合作}_{挂车}_{全部}_{所有账号}_{发布时间不限}视频详情数据.xlsx", "视频详情-合作挂车", "官方", "douyin"),
    (r"C:/Users/EDY/Downloads/[20260630-20260706]_{合作}_{非挂车}_{全部}_{所有账号}_{发布时间不限}视频详情数据.xlsx", "视频详情-合作非挂车", "官方", "douyin"),
    (r"C:/Users/EDY/Downloads/[20260701-20260731]_{自营}_{非挂车}_{全部}_{所有账号}_{发布时间不限}视频详情数据.xlsx", "视频详情-自营非挂车", "官方", "douyin"),
    (r"C:/Users/EDY/Downloads/[20260701-20260731]_{合作}_{非挂车}_{全部}_{所有账号}_{发布时间不限}视频详情数据.xlsx", "视频详情-合作非挂车", "官方", "douyin"),
    (r"C:/Users/EDY/Downloads/[20260706-20260712]_{自营}_{挂车}_{全部}_{所有账号}_{发布时间不限}视频详情数据.xlsx", "视频详情-自营挂车", "官方", "douyin"),
    (r"C:/Users/EDY/Downloads/全域数据_直播分析_弹动官方旗舰店_2026-07-09 00_00_00_2026-07-09 23_59_59_7660808418407612470.xlsx", "直播分析-全域", "官方", "douyin"),
    (r"C:/Users/EDY/Downloads/2026-08-05日直播间修正报表导出.xlsx", "直播间修正报表", "官方", "douyin"),
    (r"C:/Users/EDY/Downloads/全域数据_素材分析_视频_2026-07-09 00_00_00-2026-07-09 23_59_59-7660808418407628854.xlsx", "素材分析-视频", "官方", "douyin"),
]

HEADER = ["source_id", "file_name", "file_path", "report_name", "sheet_name", "shop", "platform",
          "file_size", "sha256", "period_start", "period_end", "download_time", "header_count",
          "row_count", "time_semantics", "candidate_grain", "candidate_business_key",
          "existing_mapping", "existing_core_object", "existing_mart_object", "status"]

def period_from_name(path, report):
    m = re.search(r"(\d{4})[-_](\d{2})[-_](\d{2})", path)
    if not m:
        m2 = re.search(r"(\d{8})-(\d{8})", path)
        if m2:
            return m2.group(1), m2.group(2)
        return "-", "-"
    # 商品卡列表格式 2026_06_30~2026_07_06
    m2 = re.search(r"(\d{4})[-_](\d{2})[-_](\d{2})~(\d{4})[-_](\d{2})[-_](\d{2})", path)
    if m2:
        return "{}-{}-{}".format(m2.group(1), m2.group(2), m2.group(3)), "{}-{}-{}".format(m2.group(4), m2.group(5), m2.group(6))
    return "{}-{}-{}".format(m.group(1), m.group(2), m.group(3)), "-"

seen = {}
out_rows = []
for idx, (path, report, shop, platform) in enumerate(CANDIDATES, 1):
    p = Path(path)
    base = [idx, p.name, path, report, "sheet1", shop, platform]
    if not p.exists():
        out_rows.append(base + ["-", "-", "-", "-", "-", "-", "-", "-", "-", "-", "-",
                                "SOURCE_NOT_AVAILABLE(本机无此文件)", "-", "-", "SOURCE_NOT_AVAILABLE"])
        continue
    h = hashlib.sha256(p.read_bytes()).hexdigest()
    dup_flag = "DUPLICATE_FILE" if h in seen else ""
    seen[h] = report
    ps, pe = period_from_name(path, report)
    try:
        wb = openpyxl.load_workbook(p, read_only=False, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows_list = list(ws.iter_rows(values_only=True))
        hdr = rows_list[0] if rows_list else ()
        ncol, nrow = len(hdr), len(rows_list) - 1
        # 时间语义粗判：文件名含精确单日→SESSION/DAILY；含区间→PERIOD_SNAPSHOT
        ts = "PERIOD_SNAPSHOT"
        if "日直播间" in report or "直播分析" in report:
            ts = "SESSION_FACT" if nrow > 0 and any("场" in str(x) for x in hdr) else "PERIOD_SNAPSHOT"
        if "视频详情" in report:
            ts = "PERIOD_SNAPSHOT"  # 文件名区间导出
        wb.close()
        out_rows.append(base + [str(p.stat().st_size), h[:16], ps, pe, "-", str(ncol), str(nrow),
                                ts, "-", "-", "-", "-", "-",
                                "DUPLICATE_FILE" if dup_flag else "DATA_ONBOARDING_GAP"])
    except Exception as e:
        out_rows.append(base + [str(p.stat().st_size), h[:16], ps, pe, "-", "-", "-",
                                "READ_ERROR:" + str(e)[:50], "-", "-", "-", "-", "-",
                                "READ_ERROR"])

out = Path(r"D:/ecommerce-data-system/workspace/docs/f1.0.3/F1.0.3_source_inventory.csv")
with out.open("w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow(HEADER)
    w.writerows(out_rows)

print("source inventory: {} 行".format(len(out_rows)))
from collections import Counter
print("状态分布:", dict(Counter(r[-1] for r in out_rows)))
for r in out_rows:
    print("  {} | {}行 {}列 | {} | {}".format(r[1][:44], r[13], r[12], r[14], r[-1]))
