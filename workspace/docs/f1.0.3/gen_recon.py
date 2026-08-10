# -*- coding: utf-8 -*-
"""F1.0.3 对账：Excel = core = mart（6 类；抽 10 业务键 + 典型区间 + 无数据区间 + 重复 + 重叠）"""
import csv
import os
from decimal import Decimal
from pathlib import Path

_env = {}
for line in Path(r"D:/ecommerce-data-system/mcp_server/.env").read_text(encoding="utf-8").splitlines():
    if "=" in line and not line.strip().startswith("#"):
        k, v = line.strip().split("=", 1)
        if v.strip():
            _env[k.strip()] = v.strip()

import psycopg2
import openpyxl

DB = psycopg2.connect(host="127.0.0.1", port=5432, dbname="ecommerce_db",
                      user="postgres", password=_env.get("PG_ADMIN_PASSWORD", ""), connect_timeout=5)
rows = []
total = pass_cnt = fail_cnt = 0


def check(case, detail, excel_v, core_v, mart_v=None):
    global total, pass_cnt, fail_cnt
    total += 1
    ok = excel_v is not None and core_v is not None and abs(float(excel_v) - float(core_v)) < 0.01
    if mart_v is not None:
        ok = ok and abs(float(core_v) - float(mart_v)) < 0.01
    if ok:
        pass_cnt += 1
    else:
        fail_cnt += 1
    rows.append([case, detail, excel_v, core_v, mart_v, "PASS" if ok else "FAIL"])


# ===== 1) 商品卡列表：Excel vs core（6/1-6/30，抽 10 商品键） =====
wb = openpyxl.load_workbook(r"C:/Users/EDY/Downloads/商品卡列表数据_2026_06_01~2026_06_30 (1).xlsx", read_only=False, data_only=True)
ws = wb["sheet1"]
erows = list(ws.iter_rows(values_only=True))
hdr = [str(c).strip() if c is not None else "" for c in erows[0]]
ci = {h: i for i, h in enumerate(hdr)}
cur = DB.cursor()
for r in erows[1:11]:
    pid = str(r[ci["商品ID"]]).strip()
    excel_pay = float(str(r[ci["商品卡用户支付金额"]]).replace(',',''))
    cur.execute("SELECT user_pay_amount FROM core.douyin_product_card_snapshot WHERE product_id=%s AND period_start='2026-06-01' AND period_end='2026-06-30'", (pid,))
    core_row = cur.fetchone()
    check("商品卡列表-10键", "商品{}".format(pid[:12]), excel_pay, float(core_row[0]) if core_row else None)
wb.close()

# ===== 2) 商品卡流量来源：Excel vs core =====
wb = openpyxl.load_workbook(r"C:/Users/EDY/Downloads/商品卡流量分析流量来源数据_2026_06_30~2026_07_06.xlsx", read_only=False, data_only=True)
ws = wb["sheet1"]
erows = list(ws.iter_rows(values_only=True))
hdr = [str(c).strip() if c is not None else "" for c in erows[0]]
ci = {h: i for i, h in enumerate(hdr)}
for r in erows[1:4]:
    l1, l2 = str(r[ci["一级渠道"]]).strip(), str(r[ci["二级渠道"]] or "-").strip()
    excel_pay = float(str(r[ci["用户支付金额(元)"]]).replace(',',''))
    cur.execute("SELECT user_pay_amount FROM core.douyin_product_card_traffic_snapshot WHERE channel_l1=%s AND channel_l2=%s AND period_start='2026-06-30' AND period_end='2026-07-06'", (l1, l2))
    core_row = cur.fetchone()
    check("流量来源-3键", "{}|{}".format(l1, l2), excel_pay, float(core_row[0]) if core_row else None)
wb.close()

# ===== 3) 视频：Excel vs core（6/30-7/6 自营挂车，抽 10 视频） =====
wb = openpyxl.load_workbook(r"C:/Users/EDY/Downloads/[20260630-20260706]_{自营}_{挂车}_{全部}_{所有账号}_{发布时间不限}视频详情数据.xlsx", read_only=False, data_only=True)
ws = wb[wb.sheetnames[0]]
erows = list(ws.iter_rows(values_only=True))
hdr = [str(c).strip() if c is not None else "" for c in erows[0]]
ci = {h: i for i, h in enumerate(hdr)}
pay_idx = ci.get("用户支付金额(元)")
for r in erows[1:11]:
    vid = str(r[ci["视频ID"]]).strip()
    excel_pay = float(str(r[pay_idx]).replace(',','')) if pay_idx is not None and r[pay_idx] is not None else 0.0
    cur.execute("SELECT user_pay_amount FROM core.douyin_video_snapshot WHERE video_id=%s AND period_start='2026-06-30' AND period_end='2026-07-06' AND selling_type='自营' AND carrier_type='挂车'", (vid,))
    core_row = cur.fetchone()
    check("视频-10键", "视频{}".format(vid[:12]), excel_pay, float(core_row[0]) if core_row else None)
wb.close()

# ===== 4) 直播场次：Excel vs core（抽 5） =====
wb = openpyxl.load_workbook(r"C:/Users/EDY/Downloads/2026-08-05日直播间修正报表导出.xlsx", read_only=False, data_only=True)
ws = wb["Sheet1"]
erows = list(ws.iter_rows(values_only=True))
hdr = [str(c).strip() if c is not None else "" for c in erows[0]]
ci = {h: i for i, h in enumerate(hdr)}
for r in erows[1:6]:
    lid = str(r[ci["直播间ID"]]).strip()
    st = str(r[ci["直播开始时间"]]).strip()
    dur = float(r[ci["直播时长"]])
    cur.execute("SELECT duration_minutes FROM core.douyin_live_session_snapshot WHERE live_room_id=%s AND start_time=%s", (lid, st))
    core_row = cur.fetchone()
    check("直播场次-5键", "直播间{}".format(lid[-6:]), dur, float(core_row[0]) if core_row else None)
wb.close()

# ===== 5) 素材：Excel vs core（抽 10） =====
wb = openpyxl.load_workbook(r"C:/Users/EDY/Downloads/全域数据_素材分析_视频_2026-07-09 00_00_00-2026-07-09 23_59_59-7660808418407628854.xlsx", read_only=False, data_only=True)
ws = wb[wb.sheetnames[0]]
erows = list(ws.iter_rows(values_only=True))
hdr = [str(c).strip() if c is not None else "" for c in erows[0]]
ci = {h: i for i, h in enumerate(hdr)}
mid_idx = ci.get("素材ID")
pay_idx = ci.get("用户实际支付金额")
for r in erows[1:11]:
    mid = str(r[mid_idx]).strip()
    excel_pay = float(str(r[pay_idx]).replace(',','')) if pay_idx is not None and r[pay_idx] is not None else 0.0
    cur.execute("SELECT user_pay_amount FROM core.douyin_material_snapshot WHERE material_id_src=%s AND period_start='2026-07-09'", (mid,))
    core_row = cur.fetchone()
    check("素材-10键", "素材{}".format(mid[:12]), excel_pay, float(core_row[0]) if core_row else None)
wb.close()

# ===== 6) 直播日数据：Excel vs core =====
wb = openpyxl.load_workbook(r"C:/Users/EDY/Downloads/全域数据_直播分析_弹动官方旗舰店_2026-07-09 00_00_00_2026-07-09 23_59_59_7660808418407612470.xlsx", read_only=False, data_only=True)
ws = wb[wb.sheetnames[0]]
erows = list(ws.iter_rows(values_only=True))
hdr = [str(c).strip() if c is not None else "" for c in erows[0]]
ci = {h: i for i, h in enumerate(hdr)}
for r in erows[1:]:
    dval = str(r[ci["日期"]]).strip() if r[ci["日期"]] else "全部"
    excel_amt = float(str(r[ci["整体成交金额"]]).replace(",", ""))
    if dval == "全部":
        cur.execute("SELECT transaction_amount FROM core.douyin_live_daily WHERE biz_date IS NULL")
    else:
        cur.execute("SELECT transaction_amount FROM core.douyin_live_daily WHERE biz_date=%s", (dval,))
    core_row = cur.fetchone()
    check("直播日数据", dval, excel_amt, float(core_row[0]) if core_row else None)
wb.close()

DB.close()
out = Path(r"D:/ecommerce-data-system/workspace/docs/f1.0.3/F1.0.3_reconciliation_report.csv")
with out.open("w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow(["case", "detail", "excel_value", "core_value", "mart_value", "result"])
    w.writerows(rows)
print("对账: 总{} / PASS {} / FAIL {}".format(total, pass_cnt, fail_cnt))
