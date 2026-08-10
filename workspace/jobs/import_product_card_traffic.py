# -*- coding: utf-8 -*-
"""F1.0.3 商品卡流量来源导入（PERIOD_SNAPSHOT：period × 渠道）"""
import hashlib
import re
import sys
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

_env = {}
for line in Path(r"D:/ecommerce-data-system/mcp_server/.env").read_text(encoding="utf-8").splitlines():
    if "=" in line and not line.strip().startswith("#"):
        k, v = line.strip().split("=", 1)
        if v.strip():
            _env[k.strip()] = v.strip()

import psycopg2
import openpyxl

HEADER_MAP = {
    "一级渠道": "channel_l1", "二级渠道": "channel_l2",
    "商品卡曝光人数": "exposure_users", "商品卡点击人数": "click_users",
    "商品卡点击率": "click_rate", "老客点击人数": "old_customer_clicks",
    "老客点击占比": "old_customer_click_share", "新客点击人数": "new_customer_clicks",
    "新客点击占比": "new_customer_click_share", "用户支付金额(元)": "user_pay_amount",
    "成交人数": "transaction_users", "成交客单价(元)": "avg_order_value",
    "点击-成交转化率": "click_to_transaction_rate",
    "首购新客数": "first_purchase_users", "首购新客占比": "first_purchase_share",
    "复购老客数": "repurchase_users", "复购老客占比": "repurchase_share",
}
NUMERIC = {v for v in HEADER_MAP.values() if v not in ("channel_l1", "channel_l2")}


def to_decimal(v):
    if v is None:
        return None
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    s = str(v).strip().replace(",", "")
    if not s or s.lower() in ("-", "--", "null", "none", "/"):
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def main():
    shop_id = int(sys.argv[1])
    fp = sys.argv[2]
    p = Path(fp)
    m = re.search(r"(\d{4})_(\d{2})_(\d{2})~(\d{4})_(\d{2})_(\d{2})", p.name)
    ps = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    pe = date(int(m.group(4)), int(m.group(5)), int(m.group(6)))
    sha = hashlib.sha256(p.read_bytes()).hexdigest()

    conn = psycopg2.connect(host="127.0.0.1", port=5432, dbname="ecommerce_db",
                            user="postgres", password=_env.get("PG_ADMIN_PASSWORD", ""), connect_timeout=5)
    cur = conn.cursor()
    cur.execute("SELECT batch_id FROM audit.import_batch WHERE file_sha256=%s AND import_status='validated'", (sha,))
    if cur.fetchone():
        print("DUPLICATE_FILE 跳过")
        conn.close()
        return

    wb = openpyxl.load_workbook(p, read_only=False, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
    unknown = [h for h in hdr if h and h not in HEADER_MAP]
    if unknown:
        print("SCHEMA_DRIFT:", unknown)
        wb.close(); conn.close(); return
    ci = {h: i for i, h in enumerate(hdr)}

    data = []
    for r in rows[1:]:
        l1 = str(r[ci["一级渠道"]]).strip() if r[ci["一级渠道"]] else ""
        if not l1 or l1.lower() in ("none", "nan", "合计", "总计"):
            continue
        rec = {"channel_l1": l1,
               "channel_l2": str(r[ci["二级渠道"]]).strip() if r[ci["二级渠道"]] else "-"}
        for h, col in HEADER_MAP.items():
            if col in NUMERIC:
                rec[col] = to_decimal(r[ci[h]])
        data.append(rec)
    wb.close()

    conn.rollback()
    cur = conn.cursor()
    cur.execute("INSERT INTO audit.import_batch(platform_code, shop_id, source_file_name, file_sha256, import_status, period_start, period_end, import_mode) "
                "VALUES('douyin',%s,%s,%s,'importing',%s,%s,'product_card_traffic') RETURNING batch_id", (shop_id, p.name, sha, ps, pe))
    batch_id = cur.fetchone()[0]
    cur.execute("DELETE FROM core.douyin_product_card_traffic_snapshot WHERE shop_id=%s AND period_start=%s AND period_end=%s", (shop_id, ps, pe))
    cols = ["shop_id", "period_start", "period_end", "channel_l1", "channel_l2", "exposure_users",
            "click_users", "click_rate", "old_customer_clicks", "old_customer_click_share",
            "new_customer_clicks", "new_customer_click_share", "user_pay_amount", "transaction_users",
            "avg_order_value", "click_to_transaction_rate", "first_purchase_users", "first_purchase_share",
            "repurchase_users", "repurchase_share", "batch_id"]
    sql = "INSERT INTO core.douyin_product_card_traffic_snapshot (" + ",".join(cols) + ") VALUES (" + ",".join(["%s"] * len(cols)) + ")"
    for rec in data:
        vals = tuple([shop_id, ps, pe, rec["channel_l1"], rec["channel_l2"]] + [rec.get(c) for c in cols[5:-1]] + [batch_id])
        cur.execute(sql, vals)
    cur.execute("SELECT count(*) FROM core.douyin_product_card_traffic_snapshot WHERE batch_id=%s", (batch_id,))
    n = cur.fetchone()[0]
    if n != len(data):
        conn.rollback()
        cur = conn.cursor()
        cur.execute("UPDATE audit.import_batch SET import_status='failed' WHERE batch_id=%s", (batch_id,))
        conn.commit()
        print("校验失败 {} vs {} → ROLLBACK".format(n, len(data)))
        conn.close(); return
    cur.execute("UPDATE audit.import_batch SET import_status='validated' WHERE batch_id=%s", (batch_id,))
    conn.commit()
    print("batch {} 写入 {} 行 ✓".format(batch_id, n))
    conn.close()


if __name__ == "__main__":
    main()
