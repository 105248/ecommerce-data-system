# -*- coding: utf-8 -*-
"""F1.0.3 直播日数据导入（DAILY_FACT：shop × date；"全部"行=周期汇总，biz_date NULL 保留）"""
import hashlib
import re
import sys
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

_env = {}
for line in Path(r"D:/ecommerce-data-system/mcp_server/.env").read_text(encoding="utf-8").splitlines():
    if "=" in line and not line.strip().startswith("#"):
        k, v = line.strip().split("=", 1)
        if v.strip():
            _env[k.strip()] = v.strip()

import psycopg2
import openpyxl

HEADER_MAP = {
    "抖音号名称": "account_name", "日期": "biz_date", "净成交ROI": "net_roi",
    "净成交金额": "net_transaction_amount", "1小时内退款率": "refund_rate_1h",
    "整体消耗": "ad_spend", "整体支付ROI": "pay_roi", "整体成交金额": "transaction_amount",
    "整体成交订单数": "transaction_orders", "整体成交订单成本": "transaction_order_cost",
    "智能优惠券金额": "smart_coupon_amount", "GPM": "gpm",
    "直播间整体曝光次数": "live_exposure_count", "直播间整体观看次数": "live_view_count",
}
NUMERIC = {v for v in HEADER_MAP.values() if v != "account_name"}


def to_decimal(v):
    if v is None:
        return None
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    s = str(v).strip().replace(",", "")
    if not s or s.lower() in ("-", "--", "null", "none", "/"):
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def main():
    shop_id = int(sys.argv[1])
    fp = sys.argv[2]
    p = Path(fp)
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", p.name)
    ps = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    pe = ps
    sha = hashlib.sha256(p.read_bytes()).hexdigest()

    conn = psycopg2.connect(host="127.0.0.1", port=5432, dbname="ecommerce_db",
                            user="postgres", password=_env.get("PG_ADMIN_PASSWORD", ""), connect_timeout=5)
    cur = conn.cursor()
    cur.execute("SELECT batch_id FROM audit.import_batch WHERE file_sha256=%s AND import_status='validated'", (sha,))
    if cur.fetchone():
        print("DUPLICATE_FILE 跳过"); conn.close(); return

    wb = openpyxl.load_workbook(p, read_only=False, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
    unknown = [h for h in hdr if h and h not in HEADER_MAP]
    if unknown:
        print("SCHEMA_DRIFT:", unknown); wb.close(); conn.close(); return
    ci = {h: i for i, h in enumerate(hdr)}

    data = []
    for r in rows[1:]:
        acct = str(r[ci["抖音号名称"]]).strip() if r[ci["抖音号名称"]] else ""
        if not acct:
            continue
        dval = r[ci["日期"]]
        dstr = str(dval).strip() if dval else ""
        biz_date = None
        if dstr and dstr != "全部":
            try:
                biz_date = date.fromisoformat(dstr)
            except ValueError:
                biz_date = None
        rec = {"account_name": acct, "biz_date": biz_date}
        for h, col in HEADER_MAP.items():
            if h in ("抖音号名称", "日期"):
                continue
            v = r[ci[h]]
            rec[col] = to_decimal(v)
        data.append(rec)
    wb.close()

    conn.rollback()
    cur = conn.cursor()
    cur.execute("INSERT INTO audit.import_batch(platform_code, shop_id, source_file_name, file_sha256, import_status, period_start, period_end, import_mode) "
                "VALUES('douyin',%s,%s,%s,'importing',%s,%s,'live_daily') RETURNING batch_id", (shop_id, p.name, sha, ps, pe))
    batch_id = cur.fetchone()[0]
    cur.execute("DELETE FROM core.douyin_live_daily WHERE shop_id=%s", (shop_id,))
    cols = ["shop_id", "biz_date", "account_name", "net_roi", "net_transaction_amount", "refund_rate_1h",
            "ad_spend", "pay_roi", "transaction_amount", "transaction_orders", "transaction_order_cost",
            "smart_coupon_amount", "gpm", "live_exposure_count", "live_view_count", "batch_id"]
    sql = "INSERT INTO core.douyin_live_daily (" + ",".join(cols) + ") VALUES (" + ",".join(["%s"] * len(cols)) + ")"
    for rec in data:
        vals = tuple([shop_id, rec["biz_date"], rec["account_name"]] + [rec.get(c) for c in cols[3:-1]] + [batch_id])
        cur.execute(sql, vals)
    cur.execute("SELECT count(*) FROM core.douyin_live_daily WHERE batch_id=%s", (batch_id,))
    n = cur.fetchone()[0]
    if n != len(data):
        conn.rollback()
        cur = conn.cursor()
        cur.execute("UPDATE audit.import_batch SET import_status='failed' WHERE batch_id=%s", (batch_id,))
        conn.commit()
        print("校验失败 {} vs {} → ROLLBACK".format(n, len(data)))
        conn.close(); return
    cur.execute("UPDATE audit.import_batch SET import_status='validated' WHERE batch_id=%s", (batch_id,))
    conn.commit()
    print("batch {} 直播日数据 {} 行 ✓".format(batch_id, n))
    conn.close()


if __name__ == "__main__":
    main()
