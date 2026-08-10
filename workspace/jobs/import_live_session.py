# -*- coding: utf-8 -*-
"""F1.0.3 直播场次导入（SESSION_FACT：统计周期×直播间×场次；不拆分钟/小时）"""
import hashlib
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path

_env = {}
for line in Path(r"D:/ecommerce-data-system/mcp_server/.env").read_text(encoding="utf-8").splitlines():
    if "=" in line and not line.strip().startswith("#"):
        k, v = line.strip().split("=", 1)
        if v.strip():
            _env[k.strip()] = v.strip()

import psycopg2
import openpyxl

HEADER_MAP = {
    "统计日期": "period_key", "直播间ID": "live_room_id", "直播间名称": "live_room_name",
    "店铺名称": "shop_name_src", "直播开始时间": "start_time", "直播结束时间": "end_time",
    "直播时长": "duration_minutes", "达人ID": "creator_id", "抖音号": "creator_uid",
    "达人昵称": "creator_nickname", "账号类型": "account_type",
}
NUMERIC = {"duration_minutes"}


def to_decimal(v):
    if v is None:
        return None
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    s = str(v).strip().replace(",", "")
    if not s or s.lower() in ("-", "--", "null", "none", "/"):
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def main():
    shop_id = int(sys.argv[1])
    fp = sys.argv[2]
    p = Path(fp)
    sha = hashlib.sha256(p.read_bytes()).hexdigest()
    conn = psycopg2.connect(host="127.0.0.1", port=5432, dbname="ecommerce_db",
                            user="postgres", password=_env.get("PG_ADMIN_PASSWORD", ""), connect_timeout=5)
    cur = conn.cursor()
    cur.execute("SELECT batch_id FROM audit.import_batch WHERE file_sha256=%s AND import_status='validated'", (sha,))
    if cur.fetchone():
        print("DUPLICATE_FILE 跳过")
        conn.close(); return

    wb = openpyxl.load_workbook(p, read_only=False, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
    unknown = [h for h in hdr if h and h not in HEADER_MAP]
    if unknown:
        print("SCHEMA_DRIFT:", unknown); wb.close(); conn.close(); return
    ci = {h: i for i, h in enumerate(hdr)}

    data = []
    for r in rows[1:]:
        lid = str(r[ci["直播间ID"]]).strip() if r[ci["直播间ID"]] else ""
        st = str(r[ci["直播开始时间"]]).strip() if r[ci["直播开始时间"]] else ""
        if not lid:
            continue
        rec = {"live_room_id": lid, "start_time": st}
        for h, col in HEADER_MAP.items():
            if h in ("直播间ID", "直播开始时间"):
                continue
            v = r[ci[h]]
            rec[col] = to_decimal(v) if col in NUMERIC else (None if v is None else str(v).strip())
        data.append(rec)
    wb.close()

    conn.rollback()
    cur = conn.cursor()
    cur.execute("INSERT INTO audit.import_batch(platform_code, shop_id, source_file_name, file_sha256, import_status, import_mode) "
                "VALUES('douyin',%s,%s,%s,'importing','live_session') RETURNING batch_id", (shop_id, p.name, sha))
    batch_id = cur.fetchone()[0]
    # SESSION_FACT 重叠：按 period_key 替换（同统计周期重导）
    cur.execute("SELECT DISTINCT period_key FROM core.douyin_live_session_snapshot")
    # 删除该文件将覆盖的所有 period_key
    periods = sorted({d["period_key"] for d in data})
    for pk in periods:
        cur.execute("DELETE FROM core.douyin_live_session_snapshot WHERE shop_id=%s AND period_key=%s", (shop_id, pk))
    cols = ["shop_id", "period_key", "live_room_id", "live_room_name", "shop_name_src",
            "start_time", "end_time", "duration_minutes", "creator_id", "creator_uid",
            "creator_nickname", "account_type", "batch_id"]
    sql = "INSERT INTO core.douyin_live_session_snapshot (" + ",".join(cols) + ") VALUES (" + ",".join(["%s"] * len(cols)) + ")"
    for rec in data:
        vals = tuple([shop_id, rec["period_key"], rec["live_room_id"], rec.get("live_room_name"),
                      rec.get("shop_name_src"), rec["start_time"], rec.get("end_time"),
                      rec.get("duration_minutes"), rec.get("creator_id"), rec.get("creator_uid"),
                      rec.get("creator_nickname"), rec.get("account_type"), batch_id])
        cur.execute(sql, vals)
    cur.execute("SELECT count(*) FROM core.douyin_live_session_snapshot WHERE batch_id=%s", (batch_id,))
    n = cur.fetchone()[0]
    if n != len(data):
        conn.rollback()
        cur = conn.cursor()
        cur.execute("UPDATE audit.import_batch SET import_status='failed' WHERE batch_id=%s", (batch_id,))
        conn.commit()
        print("校验失败 {} vs {} → ROLLBACK".format(n, len(data)))
        conn.close(); return
    cur.execute("UPDATE audit.import_batch SET import_status='validated' WHERE batch_id=%s", (batch_id,))
    conn.commit()
    print("batch {} 直播场次 {} 行 ✓".format(batch_id, n))
    conn.close()


if __name__ == "__main__":
    main()
