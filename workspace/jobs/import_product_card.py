# -*- coding: utf-8 -*-
"""F1.0.3 商品卡列表接入：Source → core（PERIOD_SNAPSHOT）
严格：按表头名映射 / Decimal / 事务内 replace / batch 记录 / 未知字段报错"""
import csv
import hashlib
import os
import re
import sys
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

_env = {}
for line in Path(r"D:/ecommerce-data-system/mcp_server/.env").read_text(encoding="utf-8").splitlines():
    if "=" in line and not line.strip().startswith("#"):
        k, v = line.strip().split("=", 1)
        if v.strip():
            _env[k.strip()] = v.strip()

import psycopg2
import openpyxl

# 表头 → 目标列（真实中文表头，逐列显式）
HEADER_MAP = {
    "商品标题": "product_title", "商品ID": "product_id", "商品链接": "product_link",
    "上架时间": "listing_time",
    "商品卡曝光人数": "exposure_users", "商品卡曝光次数": "exposure_count",
    "商品卡点击人数": "click_users", "商品卡点击次数": "click_count",
    "商品卡点击率(人数)": "click_rate_users", "商品卡点击率(次数)": "click_rate_count",
    "商品卡人均点击次数": "avg_click_per_user",
    "新客点击人数": "new_customer_clicks", "老客点击人数": "old_customer_clicks",
    "新客点击占比": "new_customer_click_share", "老客点击占比": "old_customer_click_share",
    "商品卡用户支付金额": "user_pay_amount", "商品卡成交人数": "transaction_users",
    "商品卡成交订单数": "transaction_orders", "商品卡成交客单价": "avg_order_value",
    "商品卡点击-成交转化率(人数)": "click_to_transaction_rate_users",
    "商品卡点击-成交转化率(次数)": "click_to_transaction_rate_count",
    "商品卡曝光-成交转化率(人数)": "exposure_to_transaction_rate_users",
    "商品卡曝光-成交转化率(次数)": "exposure_to_transaction_rate_count",
    "商品卡千次曝光用户支付金额": "gmv_per_1000_exposure",
    "商品卡加购人数": "add_to_cart_users", "商品卡收藏人数": "favorite_users",
    "首购用户数": "first_purchase_users", "复购用户数": "repurchase_users",
    "首购新客占比": "first_purchase_share", "复购老客占比": "repurchase_share",
    "平台扶持曝光次数": "platform_support_exposure_count",
    "销量": "sales_quantity", "销量排名": "sales_rank", "商品规格+商品价格": "spec_price_text",
    # 6/16-7/15 精简版表头别名（同语义字段改名，显式登记）
    "点击率(人数)": "click_rate_users", "点击-成交转化率(人数)": "click_to_transaction_rate_users",
    "点击率(次数)": "click_rate_count", "点击-成交转化率(次数": "click_to_transaction_rate_count",
}

NUMERIC_COLS = {v for k, v in HEADER_MAP.items() if v not in ("product_title", "product_link", "listing_time", "product_id", "spec_price_text")}


def to_decimal(v):
    if v is None:
        return None
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    s = str(v).strip().replace(",", "")
    if not s or s.lower() in ("-", "--", "null", "none", "/"):
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def period_from_name(name):
    m = re.search(r"(\d{4})_(\d{2})_(\d{2})~(\d{4})_(\d{2})_(\d{2})", name)
    if m:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3))), date(int(m.group(4)), int(m.group(5)), int(m.group(6)))
    m2 = re.search(r"(\d{4})_(\d{2})_(\d{2})~(\d{4})_(\d{2})_(\d{2})", name.replace("-", "_"))
    if m2:
        return date(int(m2.group(1)), int(m2.group(2)), int(m2.group(3))), date(int(m2.group(4)), int(m2.group(5)), int(m2.group(6)))
    raise ValueError("无法从文件名解析周期: " + name)


def main():
    shop_id = int(sys.argv[1]) if len(sys.argv) > 1 else 1  # 官方店=1
    shop_name = "弹动官方旗舰店" if shop_id == 1 else "弹动个人护理旗舰店"
    files = sys.argv[2:]
    conn = psycopg2.connect(host="127.0.0.1", port=5432, dbname="ecommerce_db",
                            user="postgres", password=_env.get("PG_ADMIN_PASSWORD", ""), connect_timeout=5)

    for fp in files:
        p = Path(fp)
        print("=== 处理 {} ===".format(p.name))
        sha = hashlib.sha256(p.read_bytes()).hexdigest()
        ps, pe = period_from_name(p.name)

        # 重复文件检查
        with conn.cursor() as cur:
            cur.execute("SELECT batch_id FROM audit.import_batch WHERE file_sha256=%s AND import_status='validated'", (sha,))
            if cur.fetchone():
                print("  DUPLICATE_FILE 跳过（已入库）")
                continue

        wb = openpyxl.load_workbook(p, read_only=False, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
        # 未知字段检查
        unknown = [h for h in hdr if h and h not in HEADER_MAP]
        if unknown:
            print("  SCHEMA_DRIFT 未知字段:", unknown[:5])
            wb.close()
            continue
        # 必需字段
        if "商品ID" not in hdr:
            print("  SCHEMA_DRIFT 缺商品ID")
            wb.close()
            continue

        col_idx = {h: i for i, h in enumerate(hdr)}
        data_rows = []
        for r in rows[1:]:
            pid = str(r[col_idx["商品ID"]]).strip() if r[col_idx["商品ID"]] is not None else ""
            if not pid or pid.lower() in ("none", "nan"):
                continue
            rec = {"product_id": pid}
            for h, col in HEADER_MAP.items():
                if h not in col_idx:
                    continue
                v = r[col_idx[h]]
                if col in NUMERIC_COLS:
                    rec[col] = to_decimal(v)
                else:
                    rec[col] = None if v is None else str(v).strip()
            data_rows.append(rec)
        wb.close()

        if not data_rows:
            print("  无有效数据行")
            continue

        # 事务：batch importing → replace 重叠 → INSERT → validated
        conn.rollback()
        with conn.cursor() as cur:
            cur.execute("INSERT INTO audit.import_batch(platform_code, shop_id, source_file_name, file_sha256, import_status, period_start, period_end, import_mode) "
                        "VALUES('douyin',%s,%s,%s,'importing',%s,%s,'product_card_snapshot') RETURNING batch_id", (shop_id, p.name, sha, ps, pe))
            batch_id = cur.fetchone()[0]

            # PERIOD_SNAPSHOT 重叠替换：同店+同区间删除
            cur.execute("DELETE FROM core.douyin_product_card_snapshot WHERE shop_id=%s AND period_start=%s AND period_end=%s",
                        (shop_id, ps, pe))

            cols = ["shop_id", "period_start", "period_end", "product_id", "product_title", "product_link",
                    "listing_time", "exposure_users", "exposure_count", "click_users", "click_count",
                    "click_rate_users", "click_rate_count", "avg_click_per_user", "new_customer_clicks",
                    "old_customer_clicks", "new_customer_click_share", "old_customer_click_share",
                    "user_pay_amount", "transaction_users", "transaction_orders", "avg_order_value",
                    "click_to_transaction_rate_users", "click_to_transaction_rate_count",
                    "exposure_to_transaction_rate_users", "exposure_to_transaction_rate_count",
                    "gmv_per_1000_exposure", "add_to_cart_users", "favorite_users", "first_purchase_users",
                    "repurchase_users", "first_purchase_share", "repurchase_share",
                    "platform_support_exposure_count", "sales_quantity", "sales_rank", "spec_price_text", "batch_id"]
            cols_sql = ",".join(cols)
            ph = ",".join(["%s"] * len(cols))
            insert_sql = "INSERT INTO core.douyin_product_card_snapshot (" + cols_sql + ") VALUES (" + ph + ")"
            for rec in data_rows:
                vals = tuple([shop_id, ps, pe] + [rec.get(c) for c in cols[3:-1]] + [batch_id])
                cur.execute(insert_sql, vals)

            # 行数/重复键校验
            cur.execute("SELECT count(*) FROM core.douyin_product_card_snapshot WHERE batch_id=%s", (batch_id,))
            n = cur.fetchone()[0]
            cur.execute("SELECT count(*) FROM (SELECT 1 FROM core.douyin_product_card_snapshot WHERE batch_id=%s "
                        "GROUP BY shop_id, period_start, period_end, product_id HAVING count(*)>1) x", (batch_id,))
            dup = cur.fetchone()[0]
            if n != len(data_rows) or dup > 0:
                conn.rollback()
                print("  校验失败 rows={} exp={} dup={} → ROLLBACK".format(n, len(data_rows), dup))
                cur = conn.cursor()
                cur.execute("UPDATE audit.import_batch SET import_status='failed' WHERE batch_id=%s", (batch_id,))
                conn.commit()
                continue
            cur.execute("UPDATE audit.import_batch SET import_status='validated' WHERE batch_id=%s", (batch_id,))
            conn.commit()
            print("  batch {} 写入 {} 行（预期 {}）✓".format(batch_id, n, len(data_rows)))
    conn.close()


if __name__ == "__main__":
    main()
