# -*- coding: utf-8 -*-
"""F1.0.3 素材分析导入（PERIOD_SNAPSHOT：period × material；源值保留不推算ROI）"""
import hashlib
import re
import sys
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

_env = {}
for line in Path(r"D:/ecommerce-data-system/mcp_server/.env").read_text(encoding="utf-8").splitlines():
    if "=" in line and not line.strip().startswith("#"):
        k, v = line.strip().split("=", 1)
        if v.strip():
            _env[k.strip()] = v.strip()

import psycopg2
import openpyxl

HEADER_MAP = {
    "素材名称": "material_name", "素材ID": "material_id_src", "素材评估": "material_evaluation",
    "素材时长": "material_duration", "素材创建时间": "material_create_time", "素材来源": "material_source",
    "标签": "tags", "整体消耗": "ad_spend", "整体展现次数": "exposure_count",
    "整体点击次数": "click_count", "整体点击率": "click_rate", "整体转化率": "conversion_rate",
    "基础消耗": "base_ad_spend", "整体支付ROI": "pay_roi", "整体成交金额": "transaction_amount",
    "整体成交订单数": "transaction_orders", "整体成交订单成本": "transaction_order_cost",
    "用户实际支付金额": "user_pay_amount", "整体千次展现费用": "cpm",
    "智能优惠券金额": "smart_coupon_amount", "电商平台补贴金额": "platform_subsidy",
    "整体未完结预售订单预估金": "pending_order_amount", "追投调控成交金额": "followup_adjust_amount",
    "追投用户实际支付金额": "followup_user_pay", "追投调控成交智能优惠券金": "followup_coupon_amount",
    "整体未完结预售订单预估金额": "pending_order_amount", "追投调控成交智能优惠券金额": "followup_coupon_amount",
    "10秒播放率": "play_10s_rate", "视频播放数": "video_play_count",
    "3秒播放率": "play_3s_rate", "视频完播率": "video_completion_rate",
}
NUMERIC = {v for v in HEADER_MAP.values() if v not in ("material_name", "material_id_src", "material_evaluation",
            "material_duration", "material_create_time", "material_source", "tags")}
NUMERIC |= {"play_10s_rate", "video_play_count", "play_3s_rate", "video_completion_rate"}


def to_decimal(v):
    if v is None:
        return None
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    s = str(v).strip().replace(",", "")
    if not s or s.lower() in ("-", "--", "null", "none", "/"):
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def main():
    shop_id = int(sys.argv[1])
    fp = sys.argv[2]
    p = Path(fp)
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})[ _](\d{2})", p.name)
    # 素材分析文件是单日（2026-07-09 00_00_00），周期=单日
    ps = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    pe = ps
    sha = hashlib.sha256(p.read_bytes()).hexdigest()

    conn = psycopg2.connect(host="127.0.0.1", port=5432, dbname="ecommerce_db",
                            user="postgres", password=_env.get("PG_ADMIN_PASSWORD", ""), connect_timeout=5)
    cur = conn.cursor()
    cur.execute("SELECT batch_id FROM audit.import_batch WHERE file_sha256=%s AND import_status='validated'", (sha,))
    if cur.fetchone():
        print("DUPLICATE_FILE 跳过"); conn.close(); return

    wb = openpyxl.load_workbook(p, read_only=False, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
    unknown = [h for h in hdr if h and h not in HEADER_MAP]
    if unknown:
        print("SCHEMA_DRIFT:", unknown[:6]); wb.close(); conn.close(); return
    if "素材ID" not in hdr:
        print("SCHEMA_DRIFT 缺素材ID"); wb.close(); conn.close(); return
    ci = {h: i for i, h in enumerate(hdr)}

    data = []
    for r in rows[1:]:
        mid = str(r[ci["素材ID"]]).strip() if r[ci["素材ID"]] else ""
        if not mid or mid.lower() in ("none", "nan"):
            continue
        rec = {"material_id_src": mid}
        for h, col in HEADER_MAP.items():
            if h == "素材ID" or h not in ci:
                continue
            v = r[ci[h]]
            rec[col] = to_decimal(v) if col in NUMERIC else (None if v is None else str(v).strip())
        data.append(rec)
    wb.close()

    conn.rollback()
    cur = conn.cursor()
    cur.execute("INSERT INTO audit.import_batch(platform_code, shop_id, source_file_name, file_sha256, import_status, period_start, period_end, import_mode) "
                "VALUES('douyin',%s,%s,%s,'importing',%s,%s,'material_snapshot') RETURNING batch_id", (shop_id, p.name, sha, ps, pe))
    batch_id = cur.fetchone()[0]
    cur.execute("DELETE FROM core.douyin_material_snapshot WHERE shop_id=%s AND period_start=%s AND period_end=%s", (shop_id, ps, pe))
    cols = ["shop_id", "period_start", "period_end", "material_id_src", "material_name", "material_evaluation",
            "material_duration", "material_create_time", "material_source", "tags", "ad_spend", "exposure_count",
            "click_count", "click_rate", "conversion_rate", "base_ad_spend", "pay_roi", "transaction_amount",
            "transaction_orders", "transaction_order_cost", "user_pay_amount", "cpm", "smart_coupon_amount",
            "platform_subsidy", "pending_order_amount", "followup_adjust_amount", "followup_user_pay",
            "followup_coupon_amount", "play_10s_rate", "video_play_count", "play_3s_rate",
            "video_completion_rate", "batch_id"]
    sql = "INSERT INTO core.douyin_material_snapshot (" + ",".join(cols) + ") VALUES (" + ",".join(["%s"] * len(cols)) + ")"
    for rec in data:
        vals = tuple([shop_id, ps, pe, rec["material_id_src"]] + [rec.get(c) for c in cols[4:-1]] + [batch_id])
        cur.execute(sql, vals)
    cur.execute("SELECT count(*) FROM core.douyin_material_snapshot WHERE batch_id=%s", (batch_id,))
    n = cur.fetchone()[0]
    if n != len(data):
        conn.rollback()
        cur = conn.cursor()
        cur.execute("UPDATE audit.import_batch SET import_status='failed' WHERE batch_id=%s", (batch_id,))
        conn.commit()
        print("校验失败 {} vs {} → ROLLBACK".format(n, len(data)))
        conn.close(); return
    cur.execute("UPDATE audit.import_batch SET import_status='validated' WHERE batch_id=%s", (batch_id,))
    conn.commit()
    print("batch {} 素材 {} 行 ✓（周期 {} ~ {}）".format(batch_id, n, ps, pe))
    conn.close()


if __name__ == "__main__":
    main()
