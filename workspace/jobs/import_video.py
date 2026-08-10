# -*- coding: utf-8 -*-
"""F1.0.3 视频详情导入（PERIOD_SNAPSHOT：period × video × 类型；按表头名映射，缺失列置NULL）"""
import hashlib
import re
import sys
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

_env = {}
for line in Path(r"D:/ecommerce-data-system/mcp_server/.env").read_text(encoding="utf-8").splitlines():
    if "=" in line and not line.strip().startswith("#"):
        k, v = line.strip().split("=", 1)
        if v.strip():
            _env[k.strip()] = v.strip()

import psycopg2
import openpyxl

HEADER_MAP = {
    "视频标题": "video_title", "视频ID": "video_id", "是否投放": "is_ad",
    "播放链接": "play_url", "发布时间": "publish_time", "达人昵称": "creator_nickname",
    "达人抖音号": "creator_uid", "带货商品ID": "product_id",
    "视频观看次数": "view_count", "用户支付金额(元)": "user_pay_amount",
    "退款金额(元)": "refund_amount", "引流直播间用户支付金额(": "live_channel_pay",
    "看后搜用户支付金额(元)": "search_after_pay", "引流店铺页用户支付金额(": "shop_page_pay",
    "点赞数": "like_count", "评论数": "comment_count", "收藏数": "favorite_count",
    "转发数": "share_count", "点击关注次数": "follow_count", "完播率": "completion_rate",
    "平均观看时长": "avg_watch_duration", "商品曝光次数": "product_exposure_count",
    "商品点击次数": "product_click_count", "成交订单数": "transaction_orders",
    "成交人数": "transaction_users",
    # 7月非挂车精简版：带(元)后缀变体
    "用户支付金额": "user_pay_amount", "退款金额": "refund_amount",
    "引流直播间用户支付金额": "live_channel_pay", "看后搜用户支付金额": "search_after_pay",
    "引流店铺页用户支付金额": "shop_page_pay",
    # 挂车版完整表头变体（(元)后缀 + 额外列，显式登记；未知列不静默忽略）
    "引流直播间用户支付金额(元)": "live_channel_pay", "引流店铺页用户支付金额(元)": "shop_page_pay",
    "退款订单数": "refund_orders",
    # 平台派生指标（显式登记，原值保留；不参与核心经营计算）
    "商品曝光点击率（次数）": "derived_exposure_click_rate",
    "商品点击成交率（次数）": "derived_click_transaction_rate",
    "商品曝光成交率（次数）": "derived_exposure_transaction_rate",
    "商品千次曝光用户支付金额(元)": "derived_gmv_per_1000_exposure",
    "直播间入口曝光次数": "live_entry_exposure_count",
    "引流直播间次数": "live_channel_entry_count",
    "引流直播成交订单数": "live_entry_transaction_orders",
    "入口千次曝光用户支付金额(元)": "live_entry_gmv_per_1000",
    "投放消耗(元)": "ad_spend",
    "看后搜商品曝光次数": "search_after_exposure_count",
    "看后搜商品点击次数": "search_after_click_count",
    "看后搜成交订单数": "search_after_transaction_orders",
    "引流其他用户支付金额(元)": "other_channel_pay",
    "引流其他成交订单数": "other_channel_transaction_orders",
    "预估佣金支出(元)": "estimated_commission",
    "引流店铺页成交订单数": "shop_page_transaction_orders",
}
NUMERIC = {v for v in HEADER_MAP.values() if v not in ("video_title", "video_id", "is_ad", "play_url",
            "publish_time", "creator_nickname", "creator_uid", "product_id")}
NUMERIC |= {"refund_orders", "derived_exposure_click_rate", "derived_click_transaction_rate",
            "derived_exposure_transaction_rate", "derived_gmv_per_1000_exposure",
            "live_entry_exposure_count", "live_channel_entry_count", "live_entry_transaction_orders",
            "live_entry_gmv_per_1000", "ad_spend", "search_after_exposure_count",
            "search_after_click_count", "search_after_transaction_orders", "other_channel_pay",
            "other_channel_transaction_orders", "estimated_commission", "shop_page_transaction_orders"}


def to_decimal(v):
    if v is None:
        return None
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    s = str(v).strip().replace(",", "")
    if not s or s.lower() in ("-", "--", "null", "none", "/"):
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def main():
    shop_id = int(sys.argv[1])
    conn = psycopg2.connect(host="127.0.0.1", port=5432, dbname="ecommerce_db",
                            user="postgres", password=_env.get("PG_ADMIN_PASSWORD", ""), connect_timeout=5)

    for fp in sys.argv[2:]:
        p = Path(fp)
        # 从文件名解析 周期/经营类型/挂车
        m = re.search(r"\[(\d{8})-(\d{8})\].*?_\{(\w+)\}_\{(\w+)\}", p.name)
        if not m:
            print("无法解析文件名:", p.name); continue
        ps = date(int(m.group(1)[:4]), int(m.group(1)[4:6]), int(m.group(1)[6:]))
        pe = date(int(m.group(2)[:4]), int(m.group(2)[4:6]), int(m.group(2)[6:]))
        selling = m.group(3)
        carrier = m.group(4)
        sha = hashlib.sha256(p.read_bytes()).hexdigest()

        cur = conn.cursor()
        cur.execute("SELECT batch_id FROM audit.import_batch WHERE file_sha256=%s AND import_status='validated'", (sha,))
        if cur.fetchone():
            print("DUPLICATE {} 跳过".format(p.name[:40]))
            conn.close(); return

        wb = openpyxl.load_workbook(p, read_only=False, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
        unknown = [h for h in hdr if h and h not in HEADER_MAP]
        if unknown:
            print("SCHEMA_DRIFT [{}]: 未知列 {}".format(p.name[:30], unknown[:4]))
            wb.close(); conn.close(); return
        if "视频ID" not in hdr:
            print("SCHEMA_DRIFT 缺视频ID"); wb.close(); conn.close(); return
        ci = {h: i for i, h in enumerate(hdr)}

        data = []
        for r in rows[1:]:
            vid = str(r[ci["视频ID"]]).strip() if r[ci["视频ID"]] else ""
            if not vid or vid.lower() in ("none", "nan"):
                continue
            rec = {"video_id": vid}
            for h, col in HEADER_MAP.items():
                if h not in ci:
                    continue
                v = r[ci[h]]
                rec[col] = to_decimal(v) if col in NUMERIC else (None if v is None else str(v).strip())
            data.append(rec)
        wb.close()

        if not data:
            print("无数据"); conn.close(); return

        conn.rollback()
        cur = conn.cursor()
        cur.execute("INSERT INTO audit.import_batch(platform_code, shop_id, source_file_name, file_sha256, import_status, period_start, period_end, import_mode) "
                    "VALUES('douyin',%s,%s,%s,'importing',%s,%s,'video_snapshot') RETURNING batch_id", (shop_id, p.name, sha, ps, pe))
        batch_id = cur.fetchone()[0]
        cur.execute("DELETE FROM core.douyin_video_snapshot WHERE shop_id=%s AND period_start=%s AND period_end=%s AND selling_type=%s AND carrier_type=%s",
                    (shop_id, ps, pe, selling, carrier))
        cols = ["shop_id", "period_start", "period_end", "selling_type", "carrier_type", "video_id",
                "video_title", "is_ad", "play_url", "publish_time", "creator_nickname", "creator_uid",
                "product_id", "view_count", "user_pay_amount", "refund_amount", "refund_orders", "live_channel_pay",
                "search_after_pay", "shop_page_pay", "like_count", "comment_count", "favorite_count",
                "share_count", "follow_count", "completion_rate", "avg_watch_duration",
                "product_exposure_count", "product_click_count", "transaction_orders", "transaction_users",
                "derived_exposure_click_rate", "derived_click_transaction_rate", "derived_exposure_transaction_rate", "derived_gmv_per_1000_exposure",
                "live_entry_exposure_count", "live_channel_entry_count", "live_entry_transaction_orders", "live_entry_gmv_per_1000",
                "ad_spend", "search_after_exposure_count", "search_after_click_count", "search_after_transaction_orders", "other_channel_pay", "other_channel_transaction_orders",
                "estimated_commission", "shop_page_transaction_orders", "batch_id"]
        sql = "INSERT INTO core.douyin_video_snapshot (" + ",".join(cols) + ") VALUES (" + ",".join(["%s"] * len(cols)) + ")"
        for rec in data:
            vals = tuple([shop_id, ps, pe, selling, carrier, rec["video_id"]] + [rec.get(c) for c in cols[6:-1]] + [batch_id])
            cur.execute(sql, vals)
        cur.execute("SELECT count(*) FROM core.douyin_video_snapshot WHERE batch_id=%s", (batch_id,))
        n = cur.fetchone()[0]
        if n != len(data):
            conn.rollback()
            cur = conn.cursor()
            cur.execute("UPDATE audit.import_batch SET import_status='failed' WHERE batch_id=%s", (batch_id,))
            conn.commit()
            print("校验失败 {} vs {} → ROLLBACK".format(n, len(data)))
            conn.close(); return
        cur.execute("UPDATE audit.import_batch SET import_status='validated' WHERE batch_id=%s", (batch_id,))
        conn.commit()
        print("[{}] {} batch {} 写入 {} 行 ✓".format(selling + carrier, p.name[:36], batch_id, n))
    conn.close()


if __name__ == "__main__":
    main()
