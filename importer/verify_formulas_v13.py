# -*- coding: utf-8 -*-
"""V1.3 公式反算核对 v2：一次性读整表，按行对齐所有字段。"""
from collections import defaultdict

from openpyxl import load_workbook

XLSX = r'D:/ecommerce-data-system/raw_files/douyin/弹动官方旗舰店/2026/2026-06/抖音电商罗盘-成交分析-20260601-20260630.xlsx'
DICT = r'C:/Users/EDY/Downloads/抖音成交分析_11工作表字段映射与正式表字典_V1.3.xlsx'

# 1. 中文 -> 英文 字段映射（按工作表）
cn2en = defaultdict(dict)
with open(r'C:/Users/EDY/WorkBuddy/2026-08-06-11-46-14/field_map.txt', encoding='utf-8') as f:
    for line in f:
        parts = line.strip().split('|')
        if len(parts) == 3:
            cn2en[parts[0].strip()][parts[1].strip()] = parts[2].strip()

# 2. V1.3 公式
wb_dict = load_workbook(DICT, data_only=True)
formulas = []
for row in wb_dict['指标公式规则'].iter_rows(min_row=2, values_only=True):
    if str(row[16]).strip() == '待首轮对账确认':
        formulas.append({
            'rule_id': row[0], 'table': row[2], 'name_cn': row[3], 'name_en': row[4],
            'formula_cn': row[7], 'numerator_expr': row[8], 'denominator': row[9],
            'multiplier': row[10] or 1,
        })

# 3. 一次性读入所有工作表（内存模式）
wb = load_workbook(XLSX, data_only=True, read_only=False)

tbl2sheet = {
    'douyin_deal_daily': '成交概览',
    'douyin_carrier_daily': '载体构成',
    'douyin_account_daily': '账号构成',
    'douyin_content_daily': '单载体构成',
    'douyin_terminal_daily': '终端构成',
    'douyin_category_daily': '品类构成',
    'douyin_product_daily': '商品构成',
    'douyin_price_band_daily': '价格带构成',
    'douyin_audience_daily': '人群构成',
}

# 预读所有工作表: sheet -> {cn_field: [values]}
sheet_data_cache = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(c).strip() if c is not None else '' for c in rows[0]]
    data_rows = rows[1:]
    colmap = {h: i for i, h in enumerate(headers) if h}
    # 每列的值列表
    cols = {}
    for h, idx in colmap.items():
        cols[h] = [r[idx] if idx < len(r) else None for r in data_rows]
    sheet_data_cache[sheet_name] = (headers, cols, len(data_rows))

def get_vals(sheet, field_en):
    """按英文字段取中文源字段的值列表。"""
    if sheet not in sheet_data_cache:
        return None
    headers, cols, n = sheet_data_cache[sheet]
    cn_field = None
    for cn, en in cn2en.get(sheet, {}).items():
        if en == field_en:
            cn_field = cn
            break
    if cn_field is None or cn_field not in cols:
        return None
    return cols[cn_field]

def to_f(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s or s in ('-', '--', 'null'):
        return None
    try:
        return float(s.replace(',', ''))
    except ValueError:
        return None

results = []
for f in formulas:
    sheet = tbl2sheet.get(f['table'])
    num_expr = f['numerator_expr']
    den = f['denominator']
    if not num_expr or not den:
        continue
    num_fields = [x.strip() for x in num_expr.replace('(', '').replace(')', '').split('+')]

    # 分子
    if len(num_fields) > 1:
        field_vals = [get_vals(sheet, nf) for nf in num_fields]
        if any(v is None for v in field_vals):
            results.append({'rule_id': f['rule_id'], 'table': f['table'], 'name': f['name_cn'],
                            'formula': f['formula_cn'], 'matched': 0, 'total': 0, 'match_rate': 0,
                            'mismatches': [{'reason': '复合分子字段缺失', 'fields': num_fields}], 'skipped': True})
            continue
        n = min(len(v) for v in field_vals)
        num_vals = [sum(to_f(field_vals[k][i]) or 0 for k in range(len(num_fields))) for i in range(n)]
    else:
        nv = get_vals(sheet, num_fields[0])
        if nv is None:
            results.append({'rule_id': f['rule_id'], 'table': f['table'], 'name': f['name_cn'],
                            'formula': f['formula_cn'], 'matched': 0, 'total': 0, 'match_rate': 0,
                            'mismatches': [{'reason': '分子字段缺失'}], 'skipped': True})
            continue
        num_vals = [to_f(v) for v in nv]

    dv = get_vals(sheet, den)
    tv = get_vals(sheet, f['name_en'])
    if dv is None or tv is None:
        results.append({'rule_id': f['rule_id'], 'table': f['table'], 'name': f['name_cn'],
                        'formula': f['formula_cn'], 'matched': 0, 'total': 0, 'match_rate': 0,
                        'mismatches': [{'reason': '分母或目标字段缺失'}], 'skipped': True})
        continue
    den_vals = [to_f(v) for v in dv]
    tgt_vals = [to_f(v) for v in tv]

    match = 0
    total = 0
    excluded = 0
    mismatches = []
    n = min(len(num_vals), len(den_vals), len(tgt_vals))
    for i in range(n):
        num = num_vals[i]
        denv = den_vals[i]
        tgt = tgt_vals[i]
        if num is None or denv is None or tgt is None:
            excluded += 1
            continue
        if num == 0 and denv == 0:
            excluded += 1
            continue
        if denv == 0:
            excluded += 1
            continue
        calc = num / denv * float(f['multiplier'])
        total += 1
        if abs(calc - tgt) < max(0.0005, abs(tgt) * 0.005):
            match += 1
        else:
            if len(mismatches) < 3:
                mismatches.append({'num': num, 'den': denv, 'calc': round(calc, 4), 'actual': tgt})
    results.append({
        'rule_id': f['rule_id'], 'table': f['table'], 'name': f['name_cn'],
        'formula': f['formula_cn'], 'matched': match, 'total': total,
        'match_rate': round(match / total * 100, 1) if total else 0,
        'mismatches': mismatches, 'excluded': excluded,
    })

# 输出
print('=== 47条待确认公式反算核对结果 ===')
print()
ok = sum(1 for r in results if r['match_rate'] >= 99)
warn = sum(1 for r in results if 50 <= r['match_rate'] < 99)
fail = sum(1 for r in results if 0 < r['match_rate'] < 50)
zero = sum(1 for r in results if r['total'] == 0)
print('公式成立(>=99%%): %d | 部分(50-99%%): %d | 不成立(0-50%%): %d | 无样本: %d' % (ok, warn, fail, zero))
print()
for r in results:
    if r['total'] == 0:
        verdict = 'NO-DATA(无有效样本)'
    elif r['match_rate'] >= 99:
        verdict = 'OK成立'
    elif r['match_rate'] >= 50:
        verdict = 'WARN部分'
    else:
        verdict = 'FAIL不成立'
    print('  [%s] %s %s: %d/%d (%.1f%%) %s' % (
        r['rule_id'], r['table'], r['name'], r['matched'], r['total'], r['match_rate'], verdict))
    if r['total'] and r['match_rate'] < 99:
        for m in r['mismatches']:
            print('       例: %s/%s=%.4f vs 实际%s' % (m.get('num'), m.get('den'), m.get('calc'), m.get('actual')))
        if r['mismatches'] and 'reason' in r['mismatches'][0]:
            print('       原因: %s' % r['mismatches'][0].get('reason'))
