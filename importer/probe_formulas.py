# -*- coding: utf-8 -*-
"""穷举 剔除退款系列费比 的正确分子/分母组合。"""
from openpyxl import load_workbook

XLSX = r'D:/ecommerce-data-system/raw_files/douyin/弹动官方旗舰店/2026/2026-06/抖音电商罗盘-成交分析-20260601-20260630.xlsx'
wb = load_workbook(XLSX, data_only=True, read_only=False)
ws = wb['载体构成']
rows = list(ws.iter_rows(values_only=True))
headers = [str(c).strip() if c is not None else '' for c in rows[0]]
data = rows[1:]

def col_vals(name):
    if name not in headers:
        return None
    i = headers.index(name)
    return [r[i] if i < len(r) else None for r in data]

def to_f(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s or s in ('-', '--', 'null'):
        return None
    try:
        return float(s.replace(',', ''))
    except ValueError:
        return None

# 候选分子字段（金额/消耗类）
num_candidates = ['投放消耗(店铺被投)', '投放消耗(店铺绑定)', '投放贡献成交金额',
                  '投放贡献成交退款金额(支付时间)', '投放贡献结算金额', '结算金额',
                  '结算金额(退款时间)', '退款金额(支付时间)', '退款金额(退款时间)',
                  '成交退款金额(支付时间)', '成交退款金额(退款时间)']
den_candidates = ['成交金额', '净成交金额', '用户支付金额', '退款后用户支付金额(支付时间)',
                  '结算金额', '结算金额(退款时间)', '投放贡献成交金额', '7日结算金额', '14日结算金额']

targets = {
    '投放费比(剔除退款、店铺被投)': ('投放消耗(店铺被投)', '投放消耗(店铺绑定)'),
    '投放费比(剔除退款、店铺绑定)': ('投放消耗(店铺绑定)', '投放消耗(店铺被投)'),
    '综合费比(剔除退款、店铺被投)': ('综合分子被投', '综合分子绑定'),
    '综合费比(剔除退款、店铺绑定)': ('综合分子绑定', '综合分子被投'),
}

def test(target, num_name, den_name, extra_sum=None):
    num = col_vals(num_name)
    den = col_vals(den_name)
    tgt = col_vals(target)
    if num is None or den is None or tgt is None:
        return None
    match = total = 0
    n = min(len(num), len(den), len(tgt))
    for i in range(n):
        nv, dv, tv = to_f(num[i]), to_f(den[i]), to_f(tgt[i])
        if nv is None or dv is None or tv is None or dv == 0 or (nv == 0 and dv == 0):
            continue
        calc = nv / dv
        total += 1
        if abs(calc - tv) < max(0.0005, abs(tv) * 0.005):
            match += 1
    if total == 0:
        return None
    return match, total, round(match / total * 100, 1)

print('=== 投放费比(剔除退款、店铺被投) 候选测试 ===')
target = '投放费比(剔除退款、店铺被投)'
results = []
for num in num_candidates:
    for den in den_candidates:
        r = test(target, num, den)
        if r:
            results.append((r[2], num, den, r))
results.sort(reverse=True)
for rate, num, den, r in results[:10]:
    print(f'  {rate}%: {num} ÷ {den} ({r[0]}/{r[1]})')

print()
print('=== 综合费比(剔除退款、店铺被投) 候选测试(分子=消耗+平台佣金+达人佣金) ===')
target2 = '综合费比(剔除退款、店铺被投)'
combos = []
for spend in ['投放消耗(店铺被投)', '投放消耗(店铺绑定)']:
    pc = col_vals('平台佣金(结算口径)')
    cc = col_vals('达人佣金(结算口径)')
    sp = col_vals(spend)
    if pc is None or cc is None or sp is None:
        continue
    n = min(len(pc), len(cc), len(sp))
    combo_vals = []
    for i in range(n):
        p, c, s = to_f(pc[i]), to_f(cc[i]), to_f(sp[i])
        combo_vals.append((p or 0) + (c or 0) + (s or 0))
    combos.append((spend, combo_vals))

for den in den_candidates:
    dv = col_vals(den)
    tv = col_vals(target2)
    if dv is None or tv is None:
        continue
    for spend, combo_vals in combos:
        match = total = 0
        n = min(len(combo_vals), len(dv), len(tv))
        for i in range(n):
            nv, dd, tt = combo_vals[i], to_f(dv[i]), to_f(tv[i])
            if nv is None or dd is None or tt is None or dd == 0 or (nv == 0 and dd == 0):
                continue
            calc = nv / dd
            total += 1
            if abs(calc - tt) < max(0.0005, abs(tt) * 0.005):
                match += 1
        if total:
            rate = round(match / total * 100, 1)
            if rate >= 50:
                print(f'  {rate}%: [{spend}+平台佣金+达人佣金] ÷ {den} ({match}/{total})')
