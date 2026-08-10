# -*- coding: utf-8 -*-
"""Excel 读取：使用 openpyxl 读取所有工作表，表头取第一行。"""
from typing import List, Optional

from openpyxl import load_workbook

from models import SheetData


class ExcelReader:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb = None

    def open(self):
        """读取模式打开，data_only=True 取计算后的值。
        注意：不用 read_only=True，因为该模式下复杂表头（隐藏列/格式）会被截断。"""
        self.wb = load_workbook(self.file_path, read_only=False, data_only=True)
        return self

    def close(self):
        if self.wb:
            self.wb.close()
            self.wb = None

    def sheet_names(self) -> List[str]:
        return list(self.wb.sheetnames)

    def read_sheet(self, sheet_name: str) -> SheetData:
        """读取一张工作表：第一行为表头，其余为数据。"""
        if sheet_name not in self.wb.sheetnames:
            return SheetData(sheet_name=sheet_name, headers=[], rows=[])

        ws = self.wb[sheet_name]
        sd = SheetData(sheet_name=sheet_name, headers=[], rows=[])

        raw_rows = []
        for row in ws.iter_rows(values_only=True):
            raw_rows.append(row)
            if len(raw_rows) == 1:
                # 表头
                sd.headers = [str(c).strip() if c is not None else "" for c in row]
            else:
                sd.raw_row_count += 1

        # 统计空白行（整行全空）
        for row in raw_rows[1:]:
            if all(c is None or str(c).strip() == "" for c in row):
                sd.blank_row_count += 1

        # 后续在 transformer 中把原始行转为映射后的 dict
        return sd

    def iter_raw_rows(self, sheet_name: str):
        """迭代原始行（不含表头），供 transformer 使用。"""
        ws = self.wb[sheet_name]
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            yield row
