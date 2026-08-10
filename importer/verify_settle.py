# -*- coding: utf-8 -*-
"""验证"剔除退款"系列正确口径: 分母=结算金额, 在3张表全量验证。"""
from openpyxl import load_workbook

XLSX = r'D:/ecommerce-data-system/raw_files/douyin/弹动官方旗舰店/2026/2026-06/抖音电商罗盘-成交分析-20260601-20260630.xlsx'
wb = load_workbook(XLSX, data_only=True, read_only=False)

sheets = ['载体构成', '账号构成', '单载体构成']
targets = {
    '投放费比(剔除退款、店铺被投)': ('投放消耗(店铺被投)', None),
    '投放费比(剔除退款、店铺绑定)': ('投放消耗(店铺绑定)', None),
    '综合费比(剔除退款、店铺被投)': ('投放消耗(店铺被投)', 'COMPOSITE'),
    '综合费比(剔除退款、店铺绑定)': ('投放消耗(店铺绑定)', 'COMPOSITE'),
}

def to_f(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s or s in ('-', '--', 'null'):
        return None
    try:
        return float(s.replace(',', ''))
    except ValueError:
        return None

print('=== 验证: 剔除退款系列 分母=结算金额 (3张表) ===')
for sheet in sheets:
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(c).strip() if c is not None else '' for c in rows[0]]
    data = rows[1:]
    if '结算金额' not in headers:
        print(f'  {sheet}: 无结算金额列, 跳过')
        continue
    print(f'  ── {sheet} ──')
    for tgt, (spend, mode) in targets.items():
        if tgt not in headers or spend not in headers:
            continue
        ti = headers.index(tgt)
        si = headers.index(spend)
        ji = headers.index('结算金额')
        pc_i = headers.index('平台佣金(结算口径)') if '平台佣金(结算口径)' in headers else None
        cc_i = headers.index('达人佣金(结算口径)') if '达人佣金(结算口径)' in headers else None
        match = total = 0
        n = len(data)
        for r in data:
            tv, sv, jv = to_f(r[ti]), to_f(r[si]), to_f(r[ji])
            if tv is None or jv is None or jv == 0:
                continue
            if mode == 'COMPOSITE':
                pv = to_f(r[pc_i]) if pc_i is not None else 0
                cv = to_f(r[cc_i]) if cc_i is not None else 0
                nv = (sv or 0) + (pv or 0) + (cv or 0)
            else:
                nv = sv
            if nv is None:
                continue
            calc = nv / jv
            total += 1
            if abs(calc - tv) < max(0.0005, abs(tv) * 0.005):
                match += 1
        rate = round(match / total * 100, 1) if total else 0
        mark = 'OK' if rate >= 99 else 'X'
        print(f'    [{mark}] {tgt}: {match}/{total} = {rate}%')
