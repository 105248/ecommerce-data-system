# -*- coding: utf-8 -*-
"""生成 V1.3 公式反算核对最终报告。"""
import json
import os
from datetime import datetime
from collections import defaultdict

from openpyxl import load_workbook

XLSX = r'D:/ecommerce-data-system/raw_files/douyin/弹动官方旗舰店/2026/2026-06/抖音电商罗盘-成交分析-20260601-20260630.xlsx'
DICT = r'C:/Users/EDY/Downloads/抖音成交分析_11工作表字段映射与正式表字典_V1.3.xlsx'
OUT = r'D:/ecommerce-data-system/logs/imports'

cn2en = defaultdict(dict)
with open(r'C:/Users/EDY/WorkBuddy/2026-08-06-11-46-14/field_map.txt', encoding='utf-8') as f:
    for line in f:
        parts = line.strip().split('|')
        if len(parts) == 3:
            cn2en[parts[0].strip()][parts[1].strip()] = parts[2].strip()

wb_dict = load_workbook(DICT, data_only=True)
formulas = []
for row in wb_dict['指标公式规则'].iter_rows(min_row=2, values_only=True):
    if str(row[16]).strip() == '待首轮对账确认':
        formulas.append({
            'rule_id': row[0], 'table': row[2], 'name_cn': row[3], 'name_en': row[4],
            'formula_cn': row[7], 'numerator_expr': row[8], 'denominator': row[9],
            'multiplier': row[10] or 1,
        })

wb = load_workbook(XLSX, data_only=True, read_only=False)
tbl2sheet = {
    'douyin_deal_daily': '成交概览', 'douyin_carrier_daily': '载体构成',
    'douyin_account_daily': '账号构成', 'douyin_content_daily': '单载体构成',
    'douyin_terminal_daily': '终端构成', 'douyin_category_daily': '品类构成',
    'douyin_product_daily': '商品构成', 'douyin_price_band_daily': '价格带构成',
    'douyin_audience_daily': '人群构成',
}
cache = {}
for sn in wb.sheetnames:
    rows = list(wb[sn].iter_rows(values_only=True))
    headers = [str(c).strip() if c is not None else '' for c in rows[0]]
    cols = {}
    for h, i in [(h, i) for i, h in enumerate(headers) if h]:
        cols[h] = [r[i] if i < len(r) else None for r in rows[1:]]
    cache[sn] = (headers, cols)

def get_vals(sheet, field_en):
    if sheet not in cache:
        return None
    headers, cols = cache[sheet]
    cn_field = next((cn for cn, en in cn2en.get(sheet, {}).items() if en == field_en), None)
    if cn_field is None or cn_field not in cols:
        return None
    return cols[cn_field]

def to_f(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s or s in ('-', '--', 'null'):
        return None
    try:
        return float(s.replace(',', ''))
    except ValueError:
        return None

results = []
for f in formulas:
    sheet = tbl2sheet.get(f['table'])
    num_fields = [x.strip() for x in (f['numerator_expr'] or '').replace('(', '').replace(')', '').split('+')]
    den = f['denominator']
    if len(num_fields) > 1:
        fvs = [get_vals(sheet, nf) for nf in num_fields]
        if any(v is None for v in fvs):
            results.append({**f, 'sheet': sheet, 'matched': 0, 'total': 0, 'match_rate': 0, 'mismatches': [], 'status': 'NO_DATA'})
            continue
        n = min(len(v) for v in fvs)
        num_vals = [sum(to_f(fvs[k][i]) or 0 for k in range(len(num_fields))) for i in range(n)]
    else:
        nv = get_vals(sheet, num_fields[0])
        num_vals = [to_f(v) for v in nv] if nv is not None else None
    dv = get_vals(sheet, den)
    tv = get_vals(sheet, f['name_en'])
    if num_vals is None or dv is None or tv is None:
        results.append({**f, 'sheet': sheet, 'matched': 0, 'total': 0, 'match_rate': 0, 'mismatches': [], 'status': 'NO_DATA'})
        continue
    den_vals = [to_f(v) for v in dv]
    tgt_vals = [to_f(v) for v in tv]
    match = total = 0
    mism = []
    n = min(len(num_vals), len(den_vals), len(tgt_vals))
    for i in range(n):
        num, denv, tgt = num_vals[i], den_vals[i], tgt_vals[i]
        if num is None or denv is None or tgt is None or denv == 0 or (num == 0 and denv == 0):
            continue
        calc = num / denv * float(f['multiplier'])
        total += 1
        if abs(calc - tgt) < max(0.0005, abs(tgt) * 0.005):
            match += 1
        elif len(mism) < 2:
            mism.append({'num': num, 'den': denv, 'calc': round(calc, 4), 'actual': tgt})
    rate = round(match / total * 100, 1) if total else 0
    if rate >= 99:
        status = '公式成立'
    elif rate >= 50:
        status = '部分成立-疑似口径问题'
    else:
        status = '不成立'
    results.append({**f, 'sheet': sheet, 'matched': match, 'total': total,
                    'match_rate': rate, 'mismatches': mism, 'status': status})

# 汇总
by_status = defaultdict(int)
for r in results:
    by_status[r['status']] += 1

report = {
    'report_type': 'V1.3公式反算核对报告',
    'generated_at': datetime.now().isoformat(),
    'data_source': '抖音电商罗盘-成交分析-20260601-20260630.xlsx',
    'checked_count': len(results),
    'summary': dict(by_status),
    'precision_issue_rules': [r['rule_id'] for r in results if r['match_rate'] >= 98 and r['match_rate'] < 99],
    'formula_issue_rules': [r['rule_id'] for r in results if r['status'] == '部分成立-疑似口径问题'],
    'details': results,
}

ts = datetime.now().strftime('%Y%m%d_%H%M%S')
json_path = os.path.join(OUT, f'{ts}_V13公式反算核对.json')
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(report, f, ensure_ascii=False, indent=2, default=str)

# TXT
lines = []
lines.append('=' * 60)
lines.append('V1.3 公式反算核对报告（47条待首轮对账确认）')
lines.append('数据源: 抖音电商罗盘-成交分析-20260601-20260630')
lines.append('=' * 60)
lines.append('【汇总】')
for k, v in by_status.items():
    lines.append(f'  {k}: {v}条')
lines.append('')
lines.append('【明细】')
for r in results:
    lines.append('  [%s] %s %s (%s): %d/%d = %.1f%% | %s' % (
        r['rule_id'], r['table'], r['name_cn'], r['name_en'], r['matched'], r['total'], r['match_rate'], r['status']))
    for m in r['mismatches']:
        lines.append('      例: %s/%s=%.4f vs 实际%s' % (m['num'], m['den'], m['calc'], m['actual']))
lines.append('')
lines.append('【精度问题规则】(公式实际成立, 仅四舍五入差异)')
lines.append('  ' + str(report['precision_issue_rules']))
lines.append('')
lines.append('【疑似口径问题规则】(公式分子/分母可能取错)')
lines.append('  ' + str(report['formula_issue_rules']))
lines.append('')
lines.append('=' * 60)
txt_path = os.path.join(OUT, f'{ts}_V13公式反算核对.txt')
with open(txt_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print('报告已生成:')
print('  JSON:', json_path)
print('  TXT :', txt_path)
print()
print('=== 汇总 ===')
for k, v in by_status.items():
    print(f'  {k}: {v}条')
print()
print('公式成立:', [r['rule_id'] for r in results if r['status'] == '公式成立'][:15], '...')
print('精度问题:', report['precision_issue_rules'])
print('疑似口径问题:', report['formula_issue_rules'])
