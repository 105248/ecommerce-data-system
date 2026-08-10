# -*- coding: utf-8 -*-
"""生成 V1.4 字典：修正12条剔除退款公式 + 47条升级已确认。"""
import shutil
from openpyxl import load_workbook

SRC = r'C:/Users/EDY/Desktop/数据库/抖音成交分析_11工作表字段映射与正式表字典_V1.3.xlsx'
DST = r'C:/Users/EDY/Desktop/数据库/抖音成交分析_11工作表字段映射与正式表字典_V1.4.xlsx'

# 待修正的 12 条剔除退款规则：rule_id -> (新分母, 新公式中文)
# 修正依据：反算验证 分母=结算金额(settlement_amount) 100%匹配
FIX_RULES = {
    19: ('settlement_amount', '投放消耗(店铺被投) ÷ 结算金额'),
    33: ('settlement_amount', '投放消耗(店铺绑定) ÷ 结算金额'),
    36: ('settlement_amount', '(投放消耗(店铺绑定)+平台佣金+达人佣金) ÷ 结算金额'),
    37: ('settlement_amount', '(投放消耗(店铺被投)+平台佣金+达人佣金) ÷ 结算金额'),
    40: ('settlement_amount', '投放消耗(店铺被投) ÷ 结算金额'),
    54: ('settlement_amount', '投放消耗(店铺绑定) ÷ 结算金额'),
    57: ('settlement_amount', '(投放消耗(店铺绑定)+平台佣金+达人佣金) ÷ 结算金额'),
    58: ('settlement_amount', '(投放消耗(店铺被投)+平台佣金+达人佣金) ÷ 结算金额'),
    61: ('settlement_amount', '投放消耗(店铺被投) ÷ 结算金额'),
    75: ('settlement_amount', '投放消耗(店铺绑定) ÷ 结算金额'),
    78: ('settlement_amount', '(投放消耗(店铺绑定)+平台佣金+达人佣金) ÷ 结算金额'),
    79: ('settlement_amount', '(投放消耗(店铺被投)+平台佣金+达人佣金) ÷ 结算金额'),
}

wb = load_workbook(SRC)
ws = wb['指标公式规则']
headers = [c.value for c in ws[1]]
# 列索引: 0=规则ID, 7=业务公式, 8=分子, 9=分母, 12=单行英文公式, 16=规则状态, 18=版本
fixed = 0
confirmed = 0
for row in ws.iter_rows(min_row=2):
    vals = [c.value for c in row]
    if vals[0] is None:
        continue
    rule_id = int(vals[0])
    # 1. 修正剔除退款公式
    if rule_id in FIX_RULES:
        new_den, new_formula = FIX_RULES[rule_id]
        row[9].value = new_den            # 分母
        row[7].value = new_formula        # 业务公式
        # 更新单行英文公式（列12）
        num_en = vals[8] if vals[8] else ''
        row[11].value = f'{num_en} / {new_den}' if num_en else None
        fixed += 1
    # 2. 47条待确认 -> 已确认
    if str(vals[16]).strip() == '待首轮对账确认':
        row[16].value = '已确认（首轮对账通过）'
        row[18].value = 'V1.4'  # 版本
        confirmed += 1

print(f'修正剔除退款公式: {fixed} 条')
print(f'升级已确认: {confirmed} 条')

# 更新 核对汇总 表
if '核对汇总' in wb.sheetnames:
    ws2 = wb['核对汇总']
    for row in ws2.iter_rows(min_row=1, max_row=30):
        for cell in row:
            if isinstance(cell.value, str) and 'V1.3' in cell.value:
                cell.value = cell.value.replace('V1.3', 'V1.4')
            if isinstance(cell.value, str) and '待首轮' in cell.value:
                cell.value = cell.value.replace('待首轮对账确认', '已确认（首轮对账通过）')
    # 更新统计
    for row in ws2.iter_rows(min_row=1, max_row=30):
        if row[0].value and '47' in str(row[0].value):
            row[0].value = '47条待确认公式全部核清（首轮对账通过，12条分母修正为结算金额）'

wb.save(DST)
print()
print('V1.4 已生成:', DST)

# 验证
wb2 = load_workbook(DST)
ws3 = wb2['指标公式规则']
fix_ok = sum(1 for row in ws3.iter_rows(min_row=2)
             if row[0].value in FIX_RULES and row[9].value == 'settlement_amount')
conf_ok = sum(1 for row in ws3.iter_rows(min_row=2)
              if str(row[16].value).strip() == '已确认（首轮对账通过）')
print('验证 - 修正公式数:', fix_ok, '| 已确认数:', conf_ok)
