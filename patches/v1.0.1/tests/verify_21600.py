# V1.0.1: 新10字段 21600值全量对账 (Excel vs core) + 列错位专项
import subprocess, os
from openpyxl import load_workbook

os.environ['PGPASSWORD'] = os.environ.get('PG_ADMIN_PASSWORD', '')
PSQL = [r'D:/pgsql16_fresh/pgsql/bin/psql.exe','-U','postgres','-h','127.0.0.1','-p','5432','-d','ecommerce_db','-t','-A','-F','|']
TMP = r'D:/ecommerce-data-system/patches/v1.0.1/tests/_tmp.sql'
def q(sql):
    with open(TMP, 'w', encoding='utf-8') as f:
        f.write(sql)
    r = subprocess.run(PSQL+['-f',TMP], capture_output=True, text=True, encoding='utf-8', errors='replace')
    if r.returncode != 0:
        print('ERR:', r.stderr[:150])
    return r.stdout

NEW = r'C:/Users/EDY/Downloads/抖音电商罗盘-成交分析-20260601-20260630 (3).xlsx'
wb = load_workbook(NEW, data_only=True, read_only=False)

def dstr(v):
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    t = str(v).strip()
    if len(t) == 8 and t.isdigit():
        return t[0:4] + '-' + t[4:6] + '-' + t[6:8]
    return t[:10]

# 10个新字段: (Excel表头, core列)
new10 = [
    ('投放消耗(店铺被投)', 'ad_spend_shop_promoted'),
    ('投放消耗(店铺绑定)', 'ad_spend_shop_bound'),
    ('投放贡献成交金额', 'ad_attributed_transaction_amount'),
    ('投放贡献成交占比', 'ad_attributed_transaction_share'),
    ('投放费比(剔除退款、店铺绑定)', 'ad_spend_rate_net_refund_shop_bound'),
    ('综合费比(剔除退款、店铺绑定)', 'total_expense_rate_net_refund_shop_bound'),
    ('投放效率(店铺被投)', 'ad_efficiency_shop_promoted'),
    ('投放效率(店铺绑定)', 'ad_efficiency_shop_bound'),
    ('全店效率(店铺被投)', 'store_efficiency_shop_promoted'),
    ('全店效率(店铺绑定)', 'store_efficiency_shop_bound'),
]

scope_map = {'成交概览': '全部', '自营成交': '自营', '合作成交': '合作'}
total_checked = 0
total_mismatch = 0
mismatch_examples = []

print('=== 新10字段 21600值全量对账 ===')
for sheet, scope in scope_map.items():
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() if c is not None else '' for c in rows[0]]
    idx_date = hdr.index('日期')
    idx_car = hdr.index('载体类型')
    idx_period = hdr.index('投放时段')
    col_idx = {h: i for i, h in enumerate(hdr)}
    sheet_checked = 0
    sheet_mismatch = 0
    for r in rows[1:]:
        d = dstr(r[idx_date])
        car = r[idx_car]
        period = r[idx_period]
        if car is None or period is None:
            continue
        # 查 core 对应行 (sale_scope + carrier + ad_period + 日期)
        core_rows = q(
            "SELECT ad_spend_shop_promoted, ad_spend_shop_bound, ad_attributed_transaction_amount, ad_attributed_transaction_share, ad_spend_rate_net_refund_shop_bound, total_expense_rate_net_refund_shop_bound, ad_efficiency_shop_promoted, ad_efficiency_shop_bound, store_efficiency_shop_promoted, store_efficiency_shop_bound FROM core.douyin_deal_daily WHERE biz_date='{}' AND sale_scope='{}' AND carrier_type='{}' AND ad_period='{}' LIMIT 1;".format(d, scope, car, period)
        ).strip()
        if not core_rows:
            print(f'  ⚠️ {sheet} {d} {scope}/{car}/{period} 在 core 未找到!')
            continue
        core_vals = core_rows.split('|')
        for i, (h, col) in enumerate(new10):
            excel_val = r[col_idx[h]] if h in col_idx else None
            core_val = core_vals[i].strip() if i < len(core_vals) else ''
            sheet_checked += 1
            if excel_val is None and core_val in ('', 'None'):
                continue  # 双NULL OK
            try:
                ev = float(excel_val) if excel_val is not None else None
                cv = float(core_val) if core_val not in ('', 'None') else None
                if ev is None and cv is None:
                    continue
                if ev is None or cv is None or abs(ev - cv) > 0.005:
                    sheet_mismatch += 1
                    if len(mismatch_examples) < 10:
                        mismatch_examples.append(f'{sheet} {d} {scope}/{car}/{period} "{h}": Excel={excel_val} core={core_val}')
            except (TypeError, ValueError):
                sheet_mismatch += 1
                if len(mismatch_examples) < 10:
                    mismatch_examples.append(f'{sheet} {d} {scope}/{car}/{period} "{h}": 解析异常 Excel={excel_val} core={core_val}')
    total_checked += sheet_checked
    total_mismatch += sheet_mismatch
    print(f'  {sheet}({scope}): 核对{sheet_checked}值 不匹配{sheet_mismatch}')

print()
print(f'=== 汇总: 核对 {total_checked} / 期望21600 | 不匹配 {total_mismatch} ===')
for e in mismatch_examples:
    print('  ❌', e)

# 列错位专项: 7个易错字段全量核对 (非新10字段的公共字段)
print()
print('=== 列错位专项: 7个公共字段全量核对 ===')
drift_cols = [
    ('商品曝光人数', 'product_exposure_user_count'),
    ('商品点击人数', 'product_click_user_count'),
    ('结算金额', 'settlement_amount'),
    ('退款金额(退款时间)', 'refund_amount_refund_time'),
    ('成交退款金额(支付时间)', 'transaction_refund_amount_pay_time'),
    ('退款率(支付时间)', 'refund_rate_pay_time'),
    ('1小时成交退款率(支付时间)', 'one_hour_refund_rate_pay_time'),
]
drift_checked = 0
drift_mismatch = 0
for sheet, scope in scope_map.items():
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() if c is not None else '' for c in rows[0]]
    idx_date = hdr.index('日期'); idx_car = hdr.index('载体类型'); idx_period = hdr.index('投放时段')
    col_idx = {h: i for i, h in enumerate(hdr)}
    for r in rows[1:]:
        d = dstr(r[idx_date]); car = r[idx_car]; period = r[idx_period]
        if car is None or period is None:
            continue
        core_rows = q(
            "SELECT product_exposure_user_count, product_click_user_count, settlement_amount, refund_amount_refund_time, transaction_refund_amount_pay_time, refund_rate_pay_time, one_hour_refund_rate_pay_time FROM core.douyin_deal_daily WHERE biz_date='{}' AND sale_scope='{}' AND carrier_type='{}' AND ad_period='{}' LIMIT 1;".format(d, scope, car, period)
        ).strip()
        if not core_rows:
            continue
        core_vals = core_rows.split('|')
        for i, (h, col) in enumerate(drift_cols):
            excel_val = r[col_idx[h]] if h in col_idx else None
            core_val = core_vals[i].strip() if i < len(core_vals) else ''
            drift_checked += 1
            try:
                ev = float(excel_val) if excel_val is not None else None
                cv = float(core_val) if core_val not in ('', 'None') else None
                if ev is None and cv is None:
                    continue
                if ev is None or cv is None or abs(ev - cv) > 0.005:
                    drift_mismatch += 1
                    if len(mismatch_examples) < 15:
                        mismatch_examples.append(f'[列错位] {sheet} {d} {scope}/{car}/{period} "{h}": Excel={excel_val} core={core_val}')
            except (TypeError, ValueError):
                drift_mismatch += 1
print(f'  列错位专项: 核对{drift_checked}值 不匹配{drift_mismatch}')
for e in mismatch_examples:
    if e.startswith('[列错位]'):
        print('  ❌', e)
