# V1.0.1: 扫描最新61列Excel 11张表 表头+行数
from openpyxl import load_workbook
import json

NEW = r'C:/Users/EDY/Downloads/抖音电商罗盘-成交分析-20260601-20260630 (3).xlsx'
wb = load_workbook(NEW, data_only=True, read_only=False)

report = {}
print('=== 11 张表: 列数/行数 ===')
total = 0
for sh in wb.sheetnames:
    ws = wb[sh]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(c).strip() if c is not None else '' for c in rows[0]]
    data_rows = len(rows) - 1
    total += data_rows
    report[sh] = {'cols': len(headers), 'rows': data_rows, 'headers': headers}
    print(f'  {sh}: {len(headers)}列 / {data_rows}行')
print(f'  总数据行: {total}')

# 重复表头检查
print()
print('=== 重复表头检查 ===')
for sh, info in report.items():
    h = info['headers']
    dup = [x for x in set(h) if h.count(x) > 1]
    if dup:
        print(f'  ❌ {sh}: 重复表头 {dup}')
    else:
        print(f'  ✅ {sh}: 无重复 ({len(h)} 唯一)')

# 保存表头JSON供后续使用
with open(r'D:/ecommerce-data-system/patches/v1.0.1/tests/new_headers.json', 'w', encoding='utf-8') as f:
    json.dump(report, f, ensure_ascii=False, indent=1)
print()
print('表头已保存: patches/v1.0.1/tests/new_headers.json')

# 前三张完整表头打印
print()
print('=== 成交概览 61列表头 ===')
for i, h in enumerate(report['成交概览']['headers'], 1):
    print(f'  {i:2d}. {h}')
