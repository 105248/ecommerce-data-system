# V1.0.1: 列错位专项 (06-05 全 72 行 x 7 字段, 批量查询)
import subprocess, os
from openpyxl import load_workbook

os.environ['PGPASSWORD'] = os.environ.get('PG_ADMIN_PASSWORD', '')
PSQL = [r'D:/pgsql16_fresh/pgsql/bin/psql.exe','-U','postgres','-h','127.0.0.1','-p','5432','-d','ecommerce_db','-t','-A','-F','|']
TMP = r'D:/ecommerce-data-system/patches/v1.0.1/tests/_tmp.sql'
def q(sql):
    with open(TMP, 'w', encoding='utf-8') as f:
        f.write(sql)
    r = subprocess.run(PSQL+['-f',TMP], capture_output=True, text=True, encoding='utf-8', errors='replace')
    return r.stdout

NEW = r'C:/Users/EDY/Downloads/抖音电商罗盘-成交分析-20260601-20260630 (3).xlsx'
wb = load_workbook(NEW, data_only=True, read_only=False)

def dstr(v):
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    t = str(v).strip()
    if len(t) == 8 and t.isdigit():
        return t[0:4] + '-' + t[4:6] + '-' + t[6:8]
    return t[:10]

drift_cols = [
    ('商品曝光人数', 'product_exposure_user_count'),
    ('商品点击人数', 'product_click_user_count'),
    ('结算金额', 'settlement_amount'),
    ('退款金额(退款时间)', 'refund_amount_refund_time'),
    ('成交退款金额(支付时间)', 'transaction_refund_amount_pay_time'),
    ('退款率(支付时间)', 'refund_rate_pay_time'),
    ('1小时成交退款率(支付时间)', 'one_hour_refund_rate_pay_time'),
]
scope_map = {'成交概览': '全部', '自营成交': '自营', '合作成交': '合作'}

# 批量取 core 06-05 全部行
core_map = {}
for scope in ['全部', '自营', '合作']:
    out = q("SELECT biz_date||'|'||carrier_type||'|'||ad_period||'|'||product_exposure_user_count||'|'||product_click_user_count||'|'||settlement_amount||'|'||refund_amount_refund_time||'|'||transaction_refund_amount_pay_time||'|'||refund_rate_pay_time||'|'||one_hour_refund_rate_pay_time FROM core.douyin_deal_daily WHERE biz_date='2026-06-05' AND sale_scope='{}';".format(scope))
    for l in out.split('\n'):
        if l.strip():
            parts = l.strip().split('|')
            core_map[(scope, parts[0], parts[1], parts[2])] = parts[3:]

checked = 0
mismatch = 0
examples = []
for sheet, scope in scope_map.items():
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() if c is not None else '' for c in rows[0]]
    idx_date = hdr.index('日期'); idx_car = hdr.index('载体类型'); idx_period = hdr.index('投放时段')
    col_idx = {h: i for i, h in enumerate(hdr)}
    for r in rows[1:]:
        d = dstr(r[idx_date])
        if d != '2026-06-05':
            continue
        car, period = r[idx_car], r[idx_period]
        if car is None or period is None:
            continue
        key = (scope, d, str(car), str(period))
        core_vals = core_map.get(key)
        if not core_vals:
            print(f'  ⚠️ core 未找到: {key}')
            continue
        for i, (h, col) in enumerate(drift_cols):
            excel_val = r[col_idx[h]] if h in col_idx else None
            core_val = core_vals[i].strip() if i < len(core_vals) else ''
            checked += 1
            try:
                ev = float(excel_val) if excel_val is not None else None
                cv = float(core_val) if core_val not in ('', 'None') else None
                if ev is None and cv is None:
                    continue
                if ev is None or cv is None or abs(ev - cv) > 0.005:
                    mismatch += 1
                    if len(examples) < 8:
                        examples.append(f'{sheet} {d} {scope}/{car}/{period} "{h}": Excel={excel_val} core={core_val}')
            except (TypeError, ValueError):
                mismatch += 1
                if len(examples) < 8:
                    examples.append(f'{sheet} {d} {scope}/{car}/{period} "{h}": 解析异常 Excel={excel_val} core={core_val}')

print(f'=== 列错位专项 (06-05 全行): 核对 {checked} 值 | 不匹配 {mismatch} ===')
for e in examples:
    print('  ❌', e)
if mismatch == 0:
    print('✅ 表头名匹配生效，无列错位 (P0 防护通过)')
